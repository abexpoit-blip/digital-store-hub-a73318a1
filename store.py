import asyncio
import sqlite3

# === DB CONNECT HELPER (busy_timeout fix) ===
def _dbc(path='/root/store.db'):
    _c = sqlite3.connect(path, timeout=15)
    try:
        _c.execute("PRAGMA busy_timeout=15000")
        _c.execute("PRAGMA journal_mode=WAL")
        _c.execute("PRAGMA synchronous=NORMAL")
    except Exception:
        pass
    return _c
# === END DB HELPER ===

import logging
import random
import sys
import uuid
import re
import os
from datetime import datetime, timedelta, timezone
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, CommandObject
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.exceptions import TelegramBadRequest
from aiogram.types import FSInputFile
from poll_handler import register_poll_handlers
from dotenv import load_dotenv
load_dotenv()

# ===== ZiniPay Auto-Deposit Helper (added by patch) =====
import requests as _zp_requests

# === [tg-ipv4-patch] force IPv4 for Telegram API (fixes ServerDisconnectedError) ===
try:
    import socket as _sock_v4
    import aiohttp as _aiohttp_v4

    if not getattr(_aiohttp_v4.TCPConnector, "_lovable_ipv4", False):
        _TCP_ORIG_INIT = _aiohttp_v4.TCPConnector.__init__

        def _tcp_ipv4_init(self, *args, **kwargs):
            kwargs.setdefault("family", _sock_v4.AF_INET)      # IPv4 only
            kwargs.setdefault("ttl_dns_cache", 300)
            kwargs.setdefault("limit", 100)
            kwargs.setdefault("enable_cleanup_closed", True)
            kwargs.setdefault("keepalive_timeout", 30)
            return _TCP_ORIG_INIT(self, *args, **kwargs)

        _aiohttp_v4.TCPConnector.__init__ = _tcp_ipv4_init
        _aiohttp_v4.TCPConnector._lovable_ipv4 = True
    print("[tg-ipv4] active (aiohttp forced to IPv4)", flush=True)
except Exception as _e_v4:  # never break the bot because of this patch
    print(f"[tg-ipv4] skipped: {_e_v4}", flush=True)
# === [/tg-ipv4-patch] ===
VPS_ADMIN_URL   = os.environ.get("VPS_ADMIN_URL", "http://localhost:3000")
DOWNLOAD_SECRET = os.environ.get("DOWNLOAD_SECRET", "")

def create_zinipay_invoice(user_id: int, username: str, amount: int):
    """admin panel এর /zinipay/create-invoice কে call করে — returns payment_url or None"""
    try:
        r = _zp_requests.post(
            f"{VPS_ADMIN_URL}/zinipay/create-invoice",
            json={"secret": DOWNLOAD_SECRET, "user_id": user_id,
                  "username": username, "amount": int(amount)},
            timeout=15,
        )
        data = r.json()
        if r.ok and data.get("ok"):
            return data.get("payment_url")
        print(f"[zinipay] create failed: {r.status_code} {data}")
        return None
    except Exception as e:
        print(f"[zinipay] exception: {e}")
        return None
# ===== /ZiniPay Helper =====

load_dotenv()


# --- CONFIGURATION ---
TOKEN = "8364765061:AAEoT6w2l74JDowWUns2EC5OT8wEcIji9Y4"
OWNER_ID = 5311644406
ADMIN_USER = "@samexpoit"
GROUP_LINK = "https://t.me/basictrick"
VPN_BOT_LINK = "https://t.me/btstoreprobot"

# Payment Info
BKASH_NUMBER = "01971814603"
NAGAD_NUMBER = "01971814603"
BINANCE_ID = "488586141"

QUOTES = [
    "“সততাই ব্যবসার মূল মূলধন।”",
    "“সেরা কোয়ালিটি ও দ্রুত ডেলিভারি আমাদের লক্ষ্য।”",
    "“আপনার সন্তুষ্টিই আমাদের সার্থকতা।”",
    "“ডিজিটাল সেবায় আমরা আছি আপনার পাশে।”"
]

BOT_VERSION = "V 10.07"

# --- REPLACEMENT DYNAMIC TIERS & TIME HELPERS ---
def get_replace_window_hours(qty: int) -> int:
    """
    Tier-based replace windows:
    1-4 pcs   : 2 hours
    5-30 pcs  : 6 hours
    31-100 pcs: 12 hours
    101+ pcs  : 24 hours
    """
    try:
        q = int(qty)
    except Exception:
        q = 1
    if q <= 4:
        return 2
    elif q <= 30:
        return 6
    elif q <= 100:
        return 12
    else:
        return 24

def get_sale_epoch(sale_id: int, date_str: str, time_str: str) -> int:
    """Return unix timestamp (seconds) of sale delivery or creation."""
    import time
    try:
        conn = _dbc()
        row = conn.execute("SELECT delivered_at FROM delivery_archive WHERE sale_id=? ORDER BY id ASC LIMIT 1", (sale_id,)).fetchone()
        conn.close()
        if row and row[0]:
            return int(row[0])
    except Exception:
        pass
    try:
        dt_str = f"{date_str} {time_str}"
        dt = datetime.strptime(dt_str, "%Y-%m-%d %I:%M %p")
        bst_tz = timezone(timedelta(hours=6))
        dt = dt.replace(tzinfo=bst_tz)
        return int(dt.timestamp())
    except Exception:
        return int(time.time())

def format_duration(seconds: int) -> str:
    """Format duration in seconds into human-readable Bengali string."""
    seconds = max(0, int(seconds))
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    if hours > 0 and minutes > 0:
        return f"{hours} ঘণ্টা {minutes} মিনিট"
    elif hours > 0:
        return f"{hours} ঘণ্টা"
    else:
        return f"{max(1, minutes)} মিনিট"

# --- VPN EMOJI MAP (Colorful & Professional) ---
VPN_EMOJIS = {
    "nord": "🛡️",
    "express": "🚀",
    "hma": "🥷",
    "ipvanish": "⚙️",
    "potato": "🥔",
    "surfshark": "🦈",
    "cyberghost": "👻",
    "bitdefender": "🛡️",
    "avast": "🔰",
    "proton": "⚛️",
    "avg": "🛡️"
}

logging.basicConfig(level=logging.INFO, stream=sys.stdout)
bot = Bot(token=TOKEN.strip())
dp = Dispatcher()
register_poll_handlers(dp)

# --- TIME HELPER ---
def is_bot_online():
    # BD Time is UTC+6
    bd_tz = timezone(timedelta(hours=6))
    now = datetime.now(bd_tz)
    hour = now.hour
    if hour >= 9 or hour < 2:
        return True
    return False

# --- BOT STATUS & MAINTENANCE HELPERS ---
_OFF_VALS = ('0', 'off', 'false', 'no', 'closed', 'disabled')

def get_bot_status():
    conn = _dbc()
    res = conn.execute("SELECT value FROM config WHERE key='bot_status'").fetchone()
    conn.close()
    return res[0] if res else 'open'

def set_bot_status(status):
    conn = _dbc()
    conn.execute("INSERT OR REPLACE INTO config (key, value) VALUES ('bot_status', ?)", (status,))
    conn.commit()
    conn.close()

def is_maintenance_mode() -> bool:
    try:
        conn = _dbc()
        row = conn.execute("SELECT value FROM config WHERE key='maintenance_mode'").fetchone()
        conn.close()
        if not row or row[0] is None:
            return False
        val = str(row[0]).strip().lower()
        return val in ('1', 'on', 'true', 'yes', 'enabled')
    except Exception:
        return False

def get_maintenance_msg() -> str:
    def_msg = (
        "🛠️ **সিস্টেম রক্ষণাবেক্ষণ চলছে**\n\n"
        "আমরা সিস্টেমের কিছু জরুরি উন্নয়নমূলক কাজ করছি।\n"
        "অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।\n\n"
        "🔐 আপনার ব্যালেন্স ও পূর্বের অর্ডার সম্পূর্ণ নিরাপদ আছে।\n"
        "🙏 ধৈর্য ধারণের জন্য ধন্যবাদ!"
    )
    try:
        conn = _dbc()
        row = conn.execute("SELECT value FROM config WHERE key='maintenance_msg'").fetchone()
        conn.close()
        if row and row[0] and str(row[0]).strip():
            return str(row[0]).strip()
    except Exception:
        pass
    return def_msg

def is_service_enabled(service_key: str) -> bool:
    """Check if a specific service (buy_service_enabled, deposit_service_enabled, etc.) is active."""
    try:
        conn = _dbc()
        row = conn.execute("SELECT value FROM config WHERE key=?", (service_key,)).fetchone()
        conn.close()
        if not row or row[0] is None:
            return True
        val = str(row[0]).strip().lower()
        return val not in _OFF_VALS
    except Exception:
        return True

# --- DATABASE SETUP ---
def init_db():
    conn = _dbc()
    cursor = conn.cursor()
    cursor.execute("PRAGMA journal_mode=WAL;")

    # Core Tables
    cursor.execute('CREATE TABLE IF NOT EXISTS stock (id INTEGER PRIMARY KEY AUTOINCREMENT, category TEXT, data TEXT)')
    cursor.execute('CREATE TABLE IF NOT EXISTS users (user_id INTEGER PRIMARY KEY, username TEXT, balance INTEGER DEFAULT 0, is_banned INTEGER DEFAULT 0)')
    cursor.execute('CREATE TABLE IF NOT EXISTS sales (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, username TEXT, category TEXT, qty INTEGER, total INTEGER, date TEXT, time TEXT)')
    cursor.execute('CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)')
    cursor.execute('CREATE TABLE IF NOT EXISTS admins (user_id INTEGER PRIMARY KEY)')

    # Payment Logs 
    cursor.execute('''CREATE TABLE IF NOT EXISTS payment_logs (req_id TEXT PRIMARY KEY)''')

    # VPN Orders Log
    cursor.execute('''CREATE TABLE IF NOT EXISTS vpn_orders (order_id TEXT PRIMARY KEY, user_id INTEGER, vpn_name TEXT, duration TEXT, price INTEGER, status TEXT, date TEXT, admin_name TEXT)''')
    
    # New Table for Support Tickets
    cursor.execute('''CREATE TABLE IF NOT EXISTS support_tickets (ticket_id TEXT PRIMARY KEY, user_id INTEGER, type TEXT, status TEXT)''')

    # Dynamic VPN Management Tables
    cursor.execute('CREATE TABLE IF NOT EXISTS vpn_brands (vpn_id TEXT PRIMARY KEY, vpn_name TEXT)')
    cursor.execute('CREATE TABLE IF NOT EXISTS vpn_packages (vpn_id TEXT, pkg_id TEXT, price INTEGER)')

    # Safe Migrations
    columns = [
        ("user_id", "INTEGER"), ("username", "TEXT"), ("amount", "INTEGER"),
        ("status", "TEXT"), ("date", "TEXT"), ("admin_name", "TEXT"), ("admin_id", "INTEGER"),
        ("timestamp", "REAL")
    ]
    for col_name, col_type in columns:
        try: cursor.execute(f"ALTER TABLE payment_logs ADD COLUMN {col_name} {col_type}")
        except: pass
        
    try: cursor.execute("ALTER TABLE sales ADD COLUMN time TEXT")
    except: pass
    
    # Safe Migrations for Support Tickets
    try: cursor.execute("ALTER TABLE support_tickets ADD COLUMN data TEXT")
    except: pass
    try: cursor.execute("ALTER TABLE support_tickets ADD COLUMN timestamp REAL")
    except: pass
    try: cursor.execute("ALTER TABLE support_tickets ADD COLUMN admin_response TEXT")
    except: pass

    # Set Default Bot Status
    cursor.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('bot_status', 'open')")

    brands_count = cursor.execute("SELECT COUNT(*) FROM vpn_brands").fetchone()[0]
    if brands_count == 0:
        default_brands = [
            ("nord", "Nord VPN"), ("express", "Express VPN"), ("hma", "HMA"),
            ("ipvanish", "IP Vanish"), ("potato", "Potato"), ("surfshark", "Surfshark"),
            ("cyberghost", "Cyberghost"), ("bitdefender", "Bitdefender"),
            ("avast", "Avast"), ("proton", "Proton"), ("avg", "AVG")
        ]
        cursor.executemany("INSERT OR IGNORE INTO vpn_brands VALUES (?, ?)", default_brands)
        default_pkgs = [
            ("nord", "7d", 50), ("nord", "9d", 60), ("nord", "14d", 90), ("nord", "30d", 150),
            ("express", "7d", 60), ("express", "30d", 180),
            ("hma", "7d", 40), ("hma", "30d", 120),
            ("ipvanish", "7d", 45), ("ipvanish", "30d", 130),
            ("potato", "7d", 30), ("potato", "30d", 100),
            ("surfshark", "7d", 55), ("surfshark", "30d", 160),
            ("cyberghost", "7d", 40), ("cyberghost", "30d", 120),
            ("bitdefender", "7d", 35), ("bitdefender", "30d", 110),
            ("avast", "7d", 35), ("avast", "30d", 110),
            ("proton", "7d", 50), ("proton", "30d", 150),
            ("avg", "7d", 35), ("avg", "30d", 110)
        ]
        cursor.executemany("INSERT OR IGNORE INTO vpn_packages VALUES (?, ?, ?)", default_pkgs)

    # Defaults Prices
    cursor.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('price_fb61', '10')")
    cursor.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('price_fb1000', '20')")
    cursor.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('price_bmig', '50')")
    cursor.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('price_bmfb', '60')")
    cursor.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('price_tempid', '15')")
    cursor.execute("INSERT OR IGNORE INTO admins (user_id) VALUES (?)", (OWNER_ID,))

    conn.commit()
    conn.close()

init_db()

# --- HELPERS ---
def get_price(category):
    conn = _dbc()
    res = conn.execute("SELECT value FROM config WHERE key=?", (f"price_{category}",)).fetchone()
    conn.close()
    return int(res[0]) if res else 20

def set_price_db(category, price):
    conn = _dbc()
    conn.execute("INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)", (f"price_{category}", str(price)))
    conn.commit()
    conn.close()

def is_admin(user_id):
    if user_id == OWNER_ID: return True
    conn = _dbc()
    res = conn.execute("SELECT user_id FROM admins WHERE user_id=?", (user_id,)).fetchone()
    conn.close()
    return bool(res)

def to_english_num(text):
    trans = str.maketrans("০১২৩৪৫৬৭৮৯", "0123456789")
    return text.translate(trans)

def format_pkg_name(pkg_id):
    if pkg_id.endswith('d'): return f"{pkg_id[:-1]} Days Package"
    if pkg_id.endswith('m'): return f"{pkg_id[:-1]} Months Package"
    if pkg_id.endswith('y'): return f"{pkg_id[:-1]} Years Package"
    return pkg_id.upper()

def get_id_by_username(input_str):
    input_str = str(input_str).strip()
    if input_str.isdigit(): return int(input_str)

    clean_name = input_str.replace("@", "").lower()
    conn = _dbc()
    query = "SELECT user_id FROM users WHERE LOWER(username) LIKE ? OR LOWER(username) = ?"
    res = conn.execute(query, (f"%{clean_name}%", f"@{clean_name}")).fetchone()
    conn.close()
    return res[0] if res else None

# --- STATES ---
class ShopStates(StatesGroup):
    waiting_for_qty = State()
    waiting_for_deposit_num = State()
    waiting_for_deposit_amount = State()
    waiting_for_screenshot = State()
    waiting_for_vpn_delivery = State()
    waiting_for_replace_data = State()
    waiting_for_complain_text = State()
    waiting_for_admin_reply = State()
    waiting_for_admin_replace = State()
    waiting_for_notice_content = State()

# --- UPDATE USER DATA ---
def get_user_data(user_id, username=None, first_name="Unknown"):
    conn = _dbc()
    cursor = conn.cursor()
    user = cursor.execute("SELECT balance, username, is_banned FROM users WHERE user_id=?", (user_id,)).fetchone()

    display_name = f"@{username}" if username else first_name

    if not user:
        cursor.execute("INSERT INTO users (user_id, username, balance, is_banned) VALUES (?, ?, 0, 0)", (user_id, display_name))
        conn.commit()
        res = (0, display_name, 0)
    else:
        if user[1] != display_name:
            cursor.execute("UPDATE users SET username=? WHERE user_id=?", (display_name, user_id))
            conn.commit()
        res = (user[0], user[1], user[2])
    conn.close()
    return res

# --- UI ---

async def show_dashboard_ui(user_id, first_name, bot_instance, chat_id):
    bal, uname, banned = get_user_data(user_id, None, first_name)
    if banned:
        await bot_instance.send_message(chat_id, "🚫 **YOU ARE BANNED** 🚫\nএই বটটি ব্যবহার করার অনুমতি আপনার নেই।")
        return

    quote = random.choice(QUOTES)
    online_status = "🟢 Online" if is_bot_online() else "🔴 Offline"
    
    dashboard = (
        f"⚡ **BASICTRICK DIGITAL STORE • {BOT_VERSION} PREMIUM** ⚡\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✨ {quote}\n\n"
        f"⏱ **ওয়ার্কিং সময়:** ৯:০০ AM - ২:০০ AM\n"
        f"🤖 **স্ট্যাটাস:** {online_status}  •  🚀 **System:** `{BOT_VERSION} Active`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 **অ্যাকাউন্ট:** {first_name}\n"
        f"🆔 **ইউজার আইডি:** `{user_id}`\n"
        f"💰 **মোট ব্যালেন্স:** `{bal}৳`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🛒 **সার্ভিস নিতে নিচের বাটন ক্লিক করুন:**"
    )

    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="🛒 আইডি কিনুন", callback_data="catalog"),
           types.InlineKeyboardButton(text="💼 BM কিনুন", callback_data="bm_catalog"))

    kb.row(types.InlineKeyboardButton(text="💳 ব্যালেন্স অ্যাড", callback_data="deposit"),
           types.InlineKeyboardButton(text="👤 প্রোফাইল", callback_data="profile"))

    kb.row(types.InlineKeyboardButton(text="🌐 VPN Services", callback_data="vpn_catalog"),
           types.InlineKeyboardButton(text="📜 Terms & Policy", callback_data="terms_policy"))

    kb.row(types.InlineKeyboardButton(text="📞 সাপোর্ট ও হেল্প", callback_data="support_menu"),
           types.InlineKeyboardButton(text="📢 কমিউনিটি গ্রুপ", url=GROUP_LINK))

    await bot_instance.send_message(chat_id, dashboard, reply_markup=kb.as_markup())

# --- ADMIN COMMANDS ---

@dp.message(Command("openbot"))
async def open_bot_cmd(message: types.Message):
    if not is_admin(message.from_user.id): return
    set_bot_status('open')
    await message.answer("✅ Bot is now **OPEN** for new users.")
    
@dp.message(Command("closebot"))
async def close_bot_cmd(message: types.Message):
    if not is_admin(message.from_user.id): return
    set_bot_status('closed')
    await message.answer("🚫 Bot is now **CLOSED** for new users. Old users can still use it.")

@dp.message(Command("replacelog"))
async def cmd_replacelog(message: types.Message):
    if not is_admin(message.from_user.id): return
    
    now_ts = datetime.now(timezone.utc).timestamp()
    day_ago = now_ts - 86400
    
    conn = _dbc()
    replaces = conn.execute("SELECT user_id, data, admin_response, timestamp FROM support_tickets WHERE type='replace' AND status='processed' AND timestamp >= ?", (day_ago,)).fetchall()
    conn.close()
    
    if not replaces:
        return await message.answer("📂 No replacements found in the last 24 hours.")
        
    filename = f"replacements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    with open(filename, "w", encoding="utf-8") as f:
        f.write("--- REPLACEMENT LOG (LAST 24 HOURS) ---\n\n")
        for r in replaces:
            dt = datetime.fromtimestamp(r[3], timezone(timedelta(hours=6))).strftime('%Y-%m-%d %I:%M %p')
            f.write(f"Time: {dt}\nUser ID: {r[0]}\nOriginal Details:\n{r[1]}\nReplaced With:\n{r[2]}\n")
            f.write("-" * 50 + "\n\n")
    
    doc = FSInputFile(filename)
    await message.answer_document(doc, caption="📁 Here is the replacement log for the last 24 hours.")
    os.remove(filename)

@dp.message(Command("addadmin"))
async def admin_add_admin(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if message.from_user.id != OWNER_ID: return
    if not command.args: return await message.answer("❌ ব্যবহার: `/addadmin 12345678`")
    try:
        new_admin_id = int(command.args.strip())
        conn = _dbc()
        conn.execute("INSERT OR IGNORE INTO admins (user_id) VALUES (?)", (new_admin_id,))
        conn.commit(); conn.close()
        await message.answer(f"✅ Admin Added: `{new_admin_id}`")
    except: await message.answer("❌ Error.")

@dp.message(Command("notice"))
async def admin_broadcast_notice_start(message: types.Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    
    kb = InlineKeyboardBuilder()
    kb.button(text="❌ Cancel", callback_data="cancel_notice")
    
    await message.answer(
        "📢 **নোটিশ সিস্টেম চালু হয়েছে**\n\n"
        "এখন আপনি যে মেসেজ, ছবি, বা ফাইল পাঠাবেন তা নোটিশ হিসেবে সেভ হবে।\n\n"
        "👉 *আপনি চাইলে আপনার গ্যালারি থেকে ছবি+ক্যাপশন পাঠাতে পারেন, অথবা অন্য চ্যানেল থেকে মেসেজ ফরোয়ার্ড করতে পারেন।*\n\n"
        "মেসেজ পাঠানো হলে নিচের Send বাটনে ক্লিক করবেন।",
        reply_markup=kb.as_markup()
    )
    await state.set_state(ShopStates.waiting_for_notice_content)


# === Delivery format helpers: Excel/Text callback stable globals ===
import time as _time_dl

import asyncio as _asyncio_dl

# === TG_TIMEOUT_RETRY_GUARD_V2 ===
import asyncio as _tg_asyncio
from aiogram import Bot as _TGBot
from aiogram.exceptions import TelegramNetworkError as _TGNetErr, TelegramRetryAfter as _TGRetryAfter

if not getattr(_TGBot, '_nx_retry_patched', False):
    _tg_orig_call = _TGBot.__call__

    async def _nx_bot_call(self, method, request_timeout=None):
        last = None
        for attempt in range(3):
            try:
                return await _tg_orig_call(self, method, request_timeout=request_timeout or 55)
            except _TGRetryAfter as e:
                last = e
                await _tg_asyncio.sleep(getattr(e, 'retry_after', 2) + 1)
            except (_TGNetErr, _tg_asyncio.TimeoutError) as e:
                last = e
                print('[tg-retry] attempt %d failed: %s' % (attempt + 1, e), flush=True)
                await _tg_asyncio.sleep(1.5 * (attempt + 1))
        raise last

    _TGBot.__call__ = _nx_bot_call
    _TGBot._nx_retry_patched = True
    print('[tg-retry] active (3x retry, timeout=55s)', flush=True)
# === END TG_TIMEOUT_RETRY_GUARD_V2 ===



_PENDING_DELIVERY = globals().get("_PENDING_DELIVERY", {})  # sale_id -> meta
_LAST_CLICK = globals().get("_LAST_CLICK", {})              # debounce

def _pending_gc():
    """Clean expired pending delivery cache safely."""
    try:
        now = _time_dl.time()
        expired = [
            sid for sid, meta in list(_PENDING_DELIVERY.items())
            if now - float(meta.get("ts", 0)) > 3600
        ]
        for sid in expired:
            _PENDING_DELIVERY.pop(sid, None)

        expired_clicks = [
            key for key, ts in list(_LAST_CLICK.items())
            if now - float(ts or 0) > 60
        ]
        for key in expired_clicks:
            _LAST_CLICK.pop(key, None)
    except Exception as e:
        print(f"[delivery] pending gc skipped: {e}")
# === /Delivery format helpers ===


# === dfmt helpers (moved up for priority) ===

def _parse_delivery_line(line):
    """Parse: 'UID PASSWORD COOKIES...' → (uid, password, cookies).
    Uses maxsplit=2 so cookies keep their spaces."""
    line = (line or "").strip()
    if not line:
        return None, None, None
    parts = line.split(None, 2)  # split on any whitespace, max 2 splits → 3 parts
    if len(parts) < 3:
        # fallback: try tab or pipe
        for sep in ("\t", "|"):
            if sep in line:
                parts = line.split(sep, 2)
                break
    while len(parts) < 3:
        parts.append("")
    return parts[0].strip(), parts[1].strip(), parts[2].strip()

def _fmt_txt_sync(items, lbl, qty):
    lines = [f"=== {lbl} × {qty} ==="]
    for idx, item in enumerate(items or [], 1):
        if isinstance(item, (tuple, list)):
            raw = item[-1] if item else ""
        else:
            raw = str(item or "")
        text = str(raw or "").strip()
        parts = text.split(None, 2)
        uid = parts[0] if len(parts) >= 1 else ""
        pw  = parts[1] if len(parts) >= 2 else ""
        ck  = parts[2] if len(parts) >= 3 else ""
        lines.append(f"\n--- #{idx} ---\nUID: {uid}\nPASS: {pw}\nCOOKIES: {ck}")
    return ("\n".join(lines)).encode("utf-8")



# === DSH XLSX DELIVERY HELPER START ===
def _fmt_xlsx_sync(items, lbl, qty):
    """Excel delivery. Accepts list of raw strings OR (sid, raw) tuples.
    Format each line as: UID PASSWORD COOKIES (cookies keep spaces)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = (str(lbl)[:31] or "Delivery")
    ws.append(["No", "UID", "PASSWORD", "COOKIES"])

    hdr_fill = PatternFill("solid", start_color="1F4E78")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center

    for idx, raw in enumerate(items or [], 1):
        # unwrap (sid, raw) tuples if present
        if isinstance(raw, (tuple, list)):
            raw = raw[-1] if raw else ""
        text = str(raw or "").strip()
        parts = text.split(None, 2)  # max 2 splits → cookies keep spaces
        uid = parts[0] if len(parts) >= 1 else ""
        pw  = parts[1] if len(parts) >= 2 else ""
        ck  = parts[2] if len(parts) >= 3 else ""
        ws.append([idx, uid, pw, ck])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 90
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# === DSH XLSX DELIVERY HELPER END ===

# ============================================================
# [DELIVERY_FORMAT_PATCH_V2] — Smart Hybrid Excel/TXT delivery
# Async file gen, zero VPS storage, Telegram CDN hosting,
# memory TTL guard, debounce, auto-fallback.
# ============================================================
_PENDING_MAX = 200          # cap memory
_PENDING_TTL = 3600         # 1 hour

# [DELIVERY_FINAL_V11]
# Canonical path: callback, ownership, archive recovery, build and upload.
# File upload uses a killable async IPv4 curl process, never aiogram multipart.
from aiogram import BaseMiddleware as _DfmtBaseMiddleware


def _dlog(msg):
    print(f"[delivery-v11] {msg}", flush=True)


def _dtoken():
    for _name in ("BOT_TOKEN", "TOKEN", "API_TOKEN", "TG_TOKEN", "BOT_API_TOKEN"):
        _value = globals().get(_name)
        if isinstance(_value, str) and ":" in _value and len(_value) > 20:
            return _value
    _value = str(getattr(bot, "token", ""))
    if ":" in _value and len(_value) > 20:
        return _value
    raise RuntimeError("bot token not found")


async def _dupload_document(chat_id, filename, payload, caption):
    # Try fast native aiogram upload first (< 400ms)
    try:
        from aiogram.types import BufferedInputFile
        _doc = BufferedInputFile(payload, filename=filename)
        await bot.send_document(chat_id=chat_id, document=_doc, caption=caption)
        return True
    except Exception as _aiogram_err:
        _dlog(f"aiogram upload fallback to curl: {_aiogram_err}")

    # Robust fallback: curl with IPv4
    import asyncio as _du_asyncio
    import json as _du_json
    import os as _du_os
    import re as _du_re
    import tempfile as _du_tempfile

    _safe = _du_re.sub(r"[^A-Za-z0-9._-]", "_", str(filename)) or "delivery.txt"
    _path = None
    _proc = None
    try:
        with _du_tempfile.NamedTemporaryFile(
                mode="wb", prefix="nx-delivery-", suffix="-" + _safe,
                dir="/tmp", delete=False) as _tmp:
            _tmp.write(payload)
            _path = _tmp.name

        # Pass URL/token through stdin so the secret is absent from ps output.
        _cmd = (
            "curl", "--ipv4", "--http1.1", "--no-keepalive",
            "--silent", "--show-error", "--fail-with-body",
            "--connect-timeout", "10", "--max-time", "50",
            "--request", "POST",
            "--form-string", f"chat_id={chat_id}",
            "--form-string", f"caption={caption or ''}",
            "--form", f"document=@{_path};filename={_safe}",
            "--config", "-",
        )
        _config = f'url = "https://api.telegram.org/bot{_dtoken()}/sendDocument"\n'.encode()
        _proc = await _du_asyncio.create_subprocess_exec(
            *_cmd, stdin=_du_asyncio.subprocess.PIPE,
            stdout=_du_asyncio.subprocess.PIPE,
            stderr=_du_asyncio.subprocess.PIPE,
        )
        try:
            _out, _err = await _du_asyncio.wait_for(_proc.communicate(_config), 55)
        except _du_asyncio.TimeoutError:
            _proc.kill()
            await _proc.wait()
            raise TimeoutError("Telegram upload killed after 55 seconds")

        _raw = (_out or b"").decode("utf-8", "replace").strip()
        _error = (_err or b"").decode("utf-8", "replace").strip()
        if _proc.returncode != 0:
            raise RuntimeError(f"curl exit={_proc.returncode}: {(_error or _raw)[:300]}")
        try:
            _result = _du_json.loads(_raw)
        except Exception as _json_error:
            raise RuntimeError(f"Telegram returned invalid JSON: {_raw[:160]}") from _json_error
        if not _result.get("ok"):
            raise RuntimeError(f"Telegram rejected file: {_result.get('description', 'unknown')}")
        return True
    finally:
        if _proc is not None and _proc.returncode is None:
            try:
                _proc.kill()
                await _proc.wait()
            except ProcessLookupError:
                pass
        if _path:
            try:
                _du_os.unlink(_path)
            except FileNotFoundError:
                pass


def _dload_archive(sale_id):
    _cn = sqlite3.connect("/root/store.db", timeout=15)
    try:
        _cn.execute("PRAGMA busy_timeout=15000")
        return _cn.execute(
            "SELECT stock_id, data, category, user_id FROM delivery_archive "
            "WHERE sale_id=? ORDER BY id ASC", (sale_id,)
        ).fetchall()
    finally:
        _cn.close()


class _DfmtDeliveryMiddleware(_DfmtBaseMiddleware):
    async def __call__(self, handler, event, data):
        import asyncio as _delivery_asyncio

        _cbdata = getattr(event, "data", "") or ""
        if not _cbdata.startswith("dfmt:"):
            return await handler(event, data)

        _sid = -1
        _dlog(f"click user={event.from_user.id} data={_cbdata}")
        try:
            _, _fmt, _sid_text = _cbdata.split(":", 2)
            if _fmt not in ("xlsx", "txt"):
                raise ValueError("invalid format")
            _sid = int(_sid_text)
        except Exception:
            await event.answer("Invalid delivery request", show_alert=True)
            return None

        try:
            try:
                await _delivery_asyncio.wait_for(
                    event.answer(f"{_fmt.upper()} ডাউনলোড হচ্ছে..."), 15)
            except Exception:
                pass

            _meta = _PENDING_DELIVERY.get(_sid)
            if not _meta:
                _rows = await _delivery_asyncio.wait_for(
                    _delivery_asyncio.to_thread(_dload_archive, _sid), 25)
                if not _rows:
                    await event.answer("Delivery data পাওয়া যায়নি। Admin কে জানান।", show_alert=True)
                    return None
                _cat = _rows[0][2] or "item"
                _label = {
                    "fb61": "FB 61", "fb1000": "FB 1000", "tempid": "Temp ID",
                    "ig": "Instagram", "fb": "Facebook", "bmig": "BM IG", "bmfb": "BM FB",
                }.get(_cat, _cat.upper())
                _meta = {
                    "user_id": _rows[0][3], "cat": _cat, "lbl": _label,
                    "qty": len(_rows), "items": [(row[0], row[1]) for row in _rows],
                }

            if _meta.get("user_id") and int(_meta["user_id"]) != int(event.from_user.id):
                await event.answer("এটা আপনার order নয়", show_alert=True)
                return None

            _items = _meta["items"]
            _label = _meta["lbl"]
            _qty = _meta["qty"]
            if _fmt == "xlsx":
                _payload = await _delivery_asyncio.wait_for(
                    _delivery_asyncio.to_thread(_fmt_xlsx_sync, _items, _label, _qty), 45)
                _filename = f"order-{_sid}-{_meta['cat']}.xlsx"
            else:
                _payload = await _delivery_asyncio.wait_for(
                    _delivery_asyncio.to_thread(_fmt_txt_sync, _items, _label, _qty), 30)
                _filename = f"order-{_sid}-{_meta['cat']}.txt"

            _dlog(f"file-ready sale={_sid} fmt={_fmt} bytes={len(_payload)}")
            _caption = f"📦 {_label} × {_qty} • Order #{_sid}\n✅ ডাউনলোড সম্পন্ন"
            await _dupload_document(event.message.chat.id, _filename, _payload, _caption)
            _dlog(f"sent sale={_sid} fmt={_fmt} bytes={len(_payload)}")

            # Note: We keep the format buttons so customer can re-download anytime!
        except Exception as _error:
            _dlog(f"error sale={_sid}: {type(_error).__name__}: {_error}")
            try:
                await _delivery_asyncio.wait_for(
                    event.message.answer("File পাঠানো যায়নি। আবার চেষ্টা করুন অথবা Admin কে জানান।"), 30)
            except Exception:
                pass
        return None


dp.callback_query.outer_middleware(_DfmtDeliveryMiddleware())
print("[delivery-v11] READY canonical middleware + async IPv4 curl uploader", flush=True)

@dp.callback_query(F.data.startswith("dfmt:"))
async def _delivery_format_cb(c: types.CallbackQuery):
    _t0 = _time_dl.time()
    try:
        _, fmt, sid_s = c.data.split(":", 2)
        sid = int(sid_s)
    except Exception:
        return await c.answer("Invalid", show_alert=True)

    # Debounce — same user + sale within 3s → ignore
    _dk = (c.from_user.id, sid)
    if _time_dl.time() - _LAST_CLICK.get(_dk, 0) < 3:
        return await c.answer("⏳ একটু অপেক্ষা করুন...")
    _LAST_CLICK[_dk] = _time_dl.time()
    if len(_LAST_CLICK) > 500: _LAST_CLICK.clear()

    _pending_gc()
    meta = _PENDING_DELIVERY.get(sid)

    # Fallback: pull from archive (works after bot restart too)
    if not meta:
        try:
            _cn = _dbc()
            _rows = _cn.execute(
                "SELECT stock_id, data, category, user_id FROM delivery_archive "
                "WHERE sale_id=? ORDER BY id ASC", (sid,)
            ).fetchall()
            _cn.close()
            if not _rows:
                return await c.answer("⚠️ Data নেই। Admin কে জানান।", show_alert=True)
            _cat = _rows[0][2] or "ITEM"
            _owner = _rows[0][3]
            _lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID",
                    "ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}                    .get(_cat, _cat.upper())
            meta = {"user_id": _owner, "cat": _cat, "lbl": _lbl,
                    "qty": len(_rows),
                    "items": [(r[0], r[1]) for r in _rows],
                    "ts": _time_dl.time()}
        except Exception as e:
            return await c.answer(f"⚠️ Load fail", show_alert=True)

    if meta.get("user_id") and meta["user_id"] != c.from_user.id:
        return await c.answer("⛔ এটা আপনার order না", show_alert=True)

    await c.answer(f"⏳ {fmt.upper()} তৈরি হচ্ছে...")

    lbl = meta["lbl"]; qty = meta["qty"]; items = meta["items"]
    fname_base = f"order-{sid}-{meta['cat']}"

    # Generate file OFF the event loop → bot stays responsive
    try:
        if fmt == "xlsx":
            try:
                data = await _asyncio_dl.to_thread(_fmt_xlsx_sync, items, lbl, qty)
                fname = f"{fname_base}.xlsx"
            except Exception as _xe:
                import traceback
                print(f"[delivery][xlsx-FAIL] sale={sid} err={type(_xe).__name__}: {_xe}")
                traceback.print_exc()
                data = await _asyncio_dl.to_thread(_fmt_txt_sync, items, lbl, qty)
                fname = f"{fname_base}.txt"
                await c.message.answer(f"⚠️ Excel unavailable ({type(_xe).__name__}) — TXT পাঠানো হলো।")
        else:
            data = await _asyncio_dl.to_thread(_fmt_txt_sync, items, lbl, qty)
            fname = f"{fname_base}.txt"
    except Exception as e:
        return await c.message.answer(f"❌ Generate fail: {e}")

    _gen_ms = int((_time_dl.time() - _t0) * 1000)

    try:
        from aiogram.types import BufferedInputFile
        await c.message.answer_document(
            BufferedInputFile(data, filename=fname),
            caption=(
                f"📦 {lbl} × {qty}  •  🆔 Order #{sid}\n"
                f"💾 Telegram-এ permanent stored — যখন খুশি re-download।"
            )
        )
        try: await c.message.edit_reply_markup(reply_markup=None)
        except Exception: pass
        # Free RAM immediately after successful send
        _PENDING_DELIVERY.pop(sid, None)
        del data
        _total_ms = int((_time_dl.time() - _t0) * 1000)
        print(f"[delivery] sale={sid} fmt={fmt} qty={qty} gen={_gen_ms}ms total={_total_ms}ms")
    except Exception as e:
        await c.message.answer(f"❌ পাঠাতে সমস্যা: {e}")

@dp.callback_query(F.data == "cancel_notice")
async def cancel_notice_action(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    await state.clear()
    try: await c.message.edit_reply_markup(reply_markup=None)
    except: pass
    await c.message.answer("✅ Notice cancelled.")

@dp.message(ShopStates.waiting_for_notice_content)
async def process_notice_content(message: types.Message, state: FSMContext):
    if message.text and message.text.startswith('/'):
        if message.text.lower() == '/cancel':
            await state.clear()
            return await message.answer("✅ Notice cancelled.")
        else:
            await state.clear()
            return

    notice_type = "text"
    file_id = None
    text = message.text or message.caption or ""
    
    if message.photo:
        notice_type = "photo"
        file_id = message.photo[-1].file_id
    elif message.document:
        notice_type = "document"
        file_id = message.document.file_id

    await state.update_data(n_type=notice_type, n_file=file_id, n_text=text)
    
    kb = InlineKeyboardBuilder()
    kb.row(
        types.InlineKeyboardButton(text="✅ Send Notice", callback_data="confirm_send_notice"),
        types.InlineKeyboardButton(text="❌ Cancel", callback_data="cancel_notice")
    )
    
    await message.reply("👆 এই মেসেজটি নোটিশ হিসেবে পাঠাতে চান?", reply_markup=kb.as_markup())

@dp.callback_query(F.data == "confirm_send_notice")
async def execute_send_notice(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    try: await c.message.edit_reply_markup(reply_markup=None)
    except: pass
    
    data = await state.get_data()
    n_type = data.get('n_type')
    n_file = data.get('n_file')
    n_text = data.get('n_text', '')
    
    formatted_text = f"┏━━━━━━━━━━━━━━━━━━━━━┓\n┣ 📢 **ADMIN NOTICE**\n┗━━━━━━━━━━━━━━━━━━━━━┛\n\n{n_text}"

    await c.message.answer("⏳ Sending Notice to all users... This might take a while.")
    
    conn = _dbc()
    users = conn.execute("SELECT user_id FROM users").fetchall()
    conn.close()

    count = 0
    for user in users:
        try:
            if n_type == 'photo':
                await bot.send_photo(user[0], n_file, caption=formatted_text, parse_mode="Markdown")
            elif n_type == 'document':
                await bot.send_document(user[0], n_file, caption=formatted_text, parse_mode="Markdown")
            else:
                await bot.send_message(user[0], formatted_text, parse_mode="Markdown")
            count += 1
            await asyncio.sleep(0.05)
        except: pass
        
    await c.message.answer(f"✅ Notice Successfully Sent to {count} users!")
    await state.clear()

@dp.message(Command("setprice"))
async def admin_set_price(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/setprice ig 50` or `/setprice tempid 15`")
    try:
        parts = command.args.split()
        input_cat, price = parts[0].lower(), int(parts[1])
        
        if input_cat == 'ig': actual_cat = 'bmig'
        elif input_cat == 'fb': actual_cat = 'bmfb'
        elif input_cat in ['fb61', 'fb1000', 'bmig', 'bmfb', 'tempid']: actual_cat = input_cat
        else: return await message.answer("❌ Use: ig, fb, fb61, fb1000, tempid")
        
        set_price_db(actual_cat, price)
        await message.answer(f"✅ Updated: {actual_cat} -> {price}৳")
    except: await message.answer("❌ Error!")

@dp.message(Command("addbalance"))
async def admin_add_bal(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/addbalance @user 100`")
    try:
        parts = command.args.split()
        target_input, amount = parts[0], int(parts[1])
        uid = get_id_by_username(target_input)
        if not uid: return await message.answer(f"❌ User not found.")
        conn = _dbc()
        conn.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (amount, uid))
        conn.commit(); conn.close()
        await bot.send_message(uid, f"🎁 **Added Balance:** {amount}৳")
        await message.answer(f"✅ Added {amount}৳ to {target_input}")
    except: await message.answer("❌ Error")

@dp.message(Command("cutbalance"))
async def admin_cut_bal(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/cutbalance @user 50`")
    try:
        parts = command.args.split()
        target_input, amount = parts[0], int(parts[1])
        uid = get_id_by_username(target_input)
        if not uid: return await message.answer(f"❌ User not found.")
        conn = _dbc()
        conn.execute("UPDATE users SET balance = balance - ? WHERE user_id = ?", (amount, uid))
        conn.commit(); conn.close()
        await bot.send_message(uid, f"⚠️ **Deducted Balance:** {amount}৳")
        await message.answer(f"✅ Cut {amount}৳ from {target_input}")
    except: await message.answer("❌ Error")

@dp.message(Command("check"))
async def admin_check_bal(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    target = command.args.strip() if command.args else ""
    if not target: return await message.answer("❌ `/check @user` or `/check all`")

    conn = _dbc()

    if target.lower() == "all":
        rows = conn.execute("SELECT username, user_id, balance FROM users WHERE balance > 0 ORDER BY balance DESC").fetchall()
        conn.close()

        if not rows: return await message.answer("📂 No users with balance.")

        report = "💰 USER BALANCES LIST:\n"
        for idx, r in enumerate(rows, 1):
            name = r[0] if r[0] else "User"
            report += f"{idx}. {name} | ID: {r[1]} | Bal: {r[2]}৳\n"

        if len(report) > 4000:
            for x in range(0, len(report), 4000):
                await message.answer(report[x:x+4000])
        else:
            await message.answer(report)
        return

    uid = get_id_by_username(target)
    if not uid:
        conn.close()
        return await message.answer("❌ User not found.")

    res = conn.execute("SELECT balance, username FROM users WHERE user_id=?", (uid,)).fetchone()
    # FETCH LAST 5 DEPOSITS
    last_deps = conn.execute("SELECT amount, date, timestamp FROM payment_logs WHERE user_id=? AND status='approved' ORDER BY timestamp DESC LIMIT 5", (uid,)).fetchall()
    conn.close()

    if res: 
        deps_list = ""

        # MANUAL_CREDIT_BLOCK
        try:
            _c2 = _dbc()
            _mc = _c2.execute(
                "SELECT delta, reason, admin_name, created_at FROM manual_credits "
                "WHERE user_id=? ORDER BY created_at DESC LIMIT 5", (uid,)).fetchall()
            _tot = _c2.execute(
                "SELECT COALESCE(SUM(delta),0), COUNT(*) FROM manual_credits WHERE user_id=?",
                (uid,)).fetchone()
            _c2.close()
        except Exception:
            _mc, _tot = [], (0, 0)
        if _mc:
            from datetime import datetime as _dt, timedelta as _td
            deps_list += "\n\n🎁 MANUAL BALANCE (Admin):\n"
            for _m in _mc:
                _ts = _m[3] or 0
                if _ts > 1e12: _ts = _ts / 1000
                _t = (_dt.utcfromtimestamp(_ts) + _td(hours=6)).strftime('%d %b %Y, %I:%M %p')
                _sign = '+' if (_m[0] or 0) >= 0 else ''
                _rsn = (' — ' + _m[1]) if _m[1] else ''
                _adm = (' [' + _m[2] + ']') if _m[2] else ''
                deps_list += f"• {_sign}{_m[0]}৳ | {_t}{_rsn}{_adm}\n"
            deps_list += f"📊 Total: {'+' if _tot[0]>=0 else ''}{_tot[0]}৳ ({_tot[1]} বার)\n"
            if _tot[1] >= 3:
                deps_list += "⚠️ এই user বারবার manual balance নিয়েছে!\n"
        else:
            deps_list += "\n\n🎁 Manual Balance: কখনো দেওয়া হয়নি\n"
        if last_deps:
            for d in last_deps:
                if len(d) > 2 and d[2]:
                    dt_str = datetime.fromtimestamp(d[2], timezone(timedelta(hours=6))).strftime('%Y-%m-%d %I:%M %p')
                else:
                    dt_str = d[1]
                deps_list += f"\n ├ {d[0]}৳ (📅 {dt_str})"
        else:
            deps_list = "\n └ No approved deposits yet."
            
        await message.answer(f"👤 Name: {res[1]}\n🆔 ID: `{uid}`\n💰 Balance: **{res[0]}৳**\n\n📥 **Last 5 Deposits:**{deps_list}")
    else: 
        await message.answer("User data missing.")

@dp.message(Command("ban"))
async def admin_ban(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/ban @user` or `/ban 12345678`")
    
    target = command.args.strip()
    uid = get_id_by_username(target)
    
    if uid:
        conn = _dbc()
        # Insert user if they are new, then set ban status
        conn.execute("INSERT OR IGNORE INTO users (user_id, username, balance, is_banned) VALUES (?, 'Unknown', 0, 0)", (uid,))
        conn.execute("UPDATE users SET is_banned = 1 WHERE user_id = ?", (uid,))
        conn.commit(); conn.close()
        await message.answer(f"🚫 Banned successfully: `{uid}`")
    else:
        await message.answer("❌ User not found.")

@dp.message(Command("unban"))
async def admin_unban(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/unban @user` or `/unban 12345678`")
    
    target = command.args.strip()
    uid = get_id_by_username(target)
    
    if uid:
        conn = _dbc()
        conn.execute("INSERT OR IGNORE INTO users (user_id, username, balance, is_banned) VALUES (?, 'Unknown', 0, 0)", (uid,))
        conn.execute("UPDATE users SET is_banned = 0 WHERE user_id = ?", (uid,))
        conn.commit(); conn.close()
        await message.answer(f"✅ Unbanned successfully: `{uid}`")
    else:
        await message.answer("❌ User not found.")

@dp.message(Command("add"))
async def admin_add_stock(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ `/add fb61 data` or `/add tempid data`")
    try:
        parts = command.args.split(maxsplit=1)
        category = parts[0].lower()
        blocks = parts[1].split("###")
        conn = _dbc(); cursor = conn.cursor()
        count = 0
        for b in blocks:
            b = b.strip()
            if not b: continue 
            
            tokens = b.split()
            if len(tokens) >= 2:
                if category == 'tempid':
                     ck = " ".join(tokens[2:]) if len(tokens) > 2 else "No Cookie"
                     formatted = f"🆔 **Temp ID:** `{tokens[0]}`\n🔑 **PASS:** `{tokens[1]}`\n🍪 **COOKIE:** `{ck}`"
                else:
                     formatted = f"🆔 **FB ID:** `{tokens[0]}`\n🔑 **PASS:** `{tokens[1]}`\n🍪 **COOKIE:** `{' '.join(tokens[2:])}`"
                
                cursor.execute("INSERT INTO stock (category, data) VALUES (?, ?)", (category, formatted))
                count += 1
        conn.commit(); conn.close()
        await message.answer(f"✅ Added {count} items to {category}")
    except: await message.answer("❌ Error")

@dp.message(Command("addbm"))
async def admin_add_bm_stock(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ `/addbm ig link`")
    try:
        parts = command.args.split(maxsplit=1)
        type_code = parts[0].lower()
        link_data = parts[1].strip()
        if type_code == "ig": cat, name = "bmig", "Instagram BM"
        elif type_code == "fb": cat, name = "bmfb", "Facebook BM"
        else: return await message.answer("❌ Type: ig or fb")
        formatted = f"🔗 **{name} Link:**\n{link_data}"
        conn = _dbc()
        conn.execute("INSERT INTO stock (category, data) VALUES (?, ?)", (cat, formatted))
        conn.commit(); conn.close()
        await message.answer(f"✅ Added 1 {name}")
    except: await message.answer("❌ Error")

@dp.message(Command("searchstock"))
async def admin_search_stock(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ ব্যবহার: `/searchstock 1000456`")
    
    search_uid = command.args.strip()
    conn = _dbc()
    items = conn.execute("SELECT id, category, data FROM stock WHERE data LIKE ?", (f"%{search_uid}%",)).fetchall()
    conn.close()
    
    if not items: return await message.answer(f"❌ '{search_uid}' নামের কোনো স্টক পাওয়া যায়নি।")
    
    msg = f"🔍 **Search Results for:** `{search_uid}`\n\n"
    for item in items:
        data_preview = item[2].replace('\n', ' ')[:80]
        msg += f"🔹 **Stock ID:** `{item[0]}` | Cat: {item[1]}\n📝 {data_preview}...\n\n"
    
    await message.answer(msg)

@dp.message(Command("viewstock"))
async def admin_view_stock(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ `/viewstock bmig`, `/viewstock ig`, or `/viewstock tempid`")

    raw_cat = command.args.strip().lower()
    if raw_cat == 'ig': cat = 'bmig'
    elif raw_cat == 'fb': cat = 'bmfb'
    else: cat = raw_cat

    await send_stock_page(message.chat.id, cat, 0)

async def send_stock_page(chat_id, cat, offset):
    conn = _dbc()
    items = conn.execute("SELECT id, data FROM stock WHERE category=? ORDER BY id DESC LIMIT 5 OFFSET ?", (cat, offset)).fetchall()
    count = conn.execute("SELECT COUNT(*) FROM stock WHERE category=?", (cat,)).fetchone()[0]
    conn.close()
    
    if not items and offset == 0: 
        return await bot.send_message(chat_id, f"❌ Empty ({cat}).")
    if not items: return
        
    msg = f"📂 **{cat} Stock ({count})** - Recent First\nDelete: `/delstock ID`\n\n"
    for item in items: 
        data_preview = item[1].replace('\n', ' ')[:50]
        msg += f"🔹 **ID:** `{item[0]}`\n📝 {data_preview}...\n\n"
        
    kb = InlineKeyboardBuilder()
    if offset > 0:
        kb.button(text="⬅️ Prev", callback_data=f"vs_{cat}_{offset-5}")
    if offset + 5 < count:
        kb.button(text="Next ➡️", callback_data=f"vs_{cat}_{offset+5}")
    
    await bot.send_message(chat_id, msg, reply_markup=kb.as_markup() if count > 5 else None)

@dp.callback_query(F.data.startswith("vs_"))
async def viewstock_paginate(c: types.CallbackQuery):
    await c.answer()
    parts = c.data.split("_")
    cat = parts[1]
    offset = int(parts[2])
    
    conn = _dbc()
    items = conn.execute("SELECT id, data FROM stock WHERE category=? ORDER BY id DESC LIMIT 5 OFFSET ?", (cat, offset)).fetchall()
    count = conn.execute("SELECT COUNT(*) FROM stock WHERE category=?", (cat,)).fetchone()[0]
    conn.close()
    
    msg = f"📂 **{cat} Stock ({count})** - Recent First\nDelete: `/delstock ID`\n\n"
    for item in items: 
        data_preview = item[1].replace('\n', ' ')[:50]
        msg += f"🔹 **ID:** `{item[0]}`\n📝 {data_preview}...\n\n"
        
    kb = InlineKeyboardBuilder()
    if offset > 0:
        kb.button(text="⬅️ Prev", callback_data=f"vs_{cat}_{offset-5}")
    if offset + 5 < count:
        kb.button(text="Next ➡️", callback_data=f"vs_{cat}_{offset+5}")
        
    await c.message.edit_text(msg, reply_markup=kb.as_markup())


@dp.message(Command("delstock"))
async def admin_del_stock(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    try:
        stock_id = int(command.args.strip())
        conn = _dbc()
        conn.execute("DELETE FROM stock WHERE id=?", (stock_id,))
        conn.commit(); conn.close()
        await message.answer(f"✅ Deleted Stock ID `{stock_id}`")
    except: await message.answer("❌ Error")

@dp.message(Command("stats"))
async def admin_stats_daily(message: types.Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    today = datetime.now().strftime("%Y-%m-%d")
    conn = _dbc()

    sales = conn.execute("SELECT username, category, qty, total, time FROM sales WHERE date=?", (today,)).fetchall()
    total = conn.execute("SELECT SUM(total) FROM sales WHERE date=?", (today,)).fetchone()[0] or 0

    dep_data = conn.execute("SELECT count(*), SUM(amount) FROM payment_logs WHERE date=? AND status='approved'", (today,)).fetchone()
    dep_count = dep_data[0] if dep_data else 0
    dep_total = dep_data[1] if dep_data and dep_data[1] else 0

    conn.close()

    report = f"📊 **TODAY ({today})**\n"
    report += f"💰 Deposits: {dep_count} Approved ({dep_total}৳)\n"
    report += f"💵 Sales: {total}৳\n\n"
    
    if sales:
        for s in sales:
            sale_time = f" 🕒 {s[4]}" if len(s) > 4 and s[4] else "" 
            report += f"👤 {s[0]} | {s[2]}x {s[1]} | {s[3]}৳{sale_time}\n"

    await message.answer(report)

@dp.message(Command("stats_week"))
async def admin_stats_weekly(message: types.Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return

    try:
        conn = _dbc()
        report = "📊 **LAST 7 DAYS REPORT**\n━━━━━━━━━━━━━━━━\n"
        grand_total_sales = 0
        grand_total_deposits = 0

        for i in range(6, -1, -1):
            date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")

            s_row = conn.execute("SELECT SUM(total) FROM sales WHERE date=?", (date,)).fetchone()
            day_sales = s_row[0] if s_row and s_row[0] else 0

            d_row = conn.execute("SELECT SUM(amount) FROM payment_logs WHERE date=? AND status='approved'", (date,)).fetchone()
            day_deps = d_row[0] if d_row and d_row[0] else 0

            report += f"📅 {date}\n   💰 Dep: {day_deps}৳ | 🛒 Sale: {day_sales}৳\n"
            grand_total_sales += day_sales
            grand_total_deposits += day_deps

        report += f"\n━━━━━━━━━━━━━━━━\n💵 **Total Sales:** {grand_total_sales}৳\n💰 **Total Deposits:** {grand_total_deposits}৳"

        await message.answer(report)
        conn.close()

    except Exception as e:
        await message.answer(f"❌ Error computing weekly stats: {e}")

# --- VPN ADMIN COMMANDS ---

@dp.message(Command("listvpn"))
async def admin_list_vpn(message: types.Message, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    
    conn = _dbc()
    brands = conn.execute("SELECT vpn_id, vpn_name FROM vpn_brands").fetchall()
    
    if not brands:
        conn.close()
        return await message.answer("📂 No VPNs found. Use `/addvpn` to add.")
        
    msg = "🌐 **All VPN Brands & Packages:**\n━━━━━━━━━━━━━━━━━━━━\n"
    for v_id, v_name in brands:
        pkgs = conn.execute("SELECT pkg_id, price FROM vpn_packages WHERE vpn_id=?", (v_id,)).fetchall()
        emoji = VPN_EMOJIS.get(v_id, "⚛️")
        msg += f"{emoji} **{v_name}**\n"
        msg += f"🆔 **VPN ID:** `{v_id}`\n"
        
        if pkgs:
            for p in pkgs:
                msg += f"  ├ 📦 `{p[0]}`: {p[1]}৳\n"
        else:
            msg += "  └ ❌ No packages.\n"
        msg += "\n"
        
    conn.close()
    
    if len(msg) > 4000:
        for x in range(0, len(msg), 4000):
            await message.answer(msg[x:x+4000], parse_mode="Markdown")
    else:
        await message.answer(msg, parse_mode="Markdown")

@dp.message(Command("addvpn"))
async def admin_add_vpn(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/addvpn nord Nord VPN`")
    parts = command.args.split(maxsplit=1)
    if len(parts) < 2: return await message.answer("❌ Error format.")
    conn = _dbc()
    conn.execute("INSERT OR REPLACE INTO vpn_brands (vpn_id, vpn_name) VALUES (?, ?)", (parts[0].lower(), parts[1]))
    conn.commit(); conn.close()
    await message.answer(f"✅ Added VPN Brand: {parts[1]} ({parts[0].lower()})")

@dp.message(Command("delvpn"))
async def admin_del_vpn(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/delvpn nord`")
    vid = command.args.strip().lower()
    conn = _dbc()
    conn.execute("DELETE FROM vpn_brands WHERE vpn_id=?", (vid,))
    conn.execute("DELETE FROM vpn_packages WHERE vpn_id=?", (vid,))
    conn.commit(); conn.close()
    await message.answer(f"✅ Deleted VPN & Packages for: {vid}")

@dp.message(Command("delvpnprice"))
async def admin_del_vpn_price(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/delvpnprice nord 7d`")
    parts = command.args.split()
    if len(parts) < 2: return await message.answer("❌ Format Error.")
    vid, pkg = parts[0].lower(), parts[1].lower()
    conn = _dbc()
    conn.execute("DELETE FROM vpn_packages WHERE vpn_id=? AND pkg_id=?", (vid, pkg))
    conn.commit()
    conn.close()
    await message.answer(f"✅ Deleted package `{pkg}` from VPN `{vid}`")

@dp.message(Command("setvpnprice"))
async def admin_set_vpn_price(message: types.Message, command: CommandObject, state: FSMContext):
    await state.clear()
    if not is_admin(message.from_user.id): return
    if not command.args: return await message.answer("❌ Format: `/setvpnprice nord 14d 80`")
    try:
        parts = command.args.split()
        vid, pkg, price = parts[0].lower(), parts[1].lower(), int(parts[2])
        conn = _dbc()
        exists = conn.execute("SELECT 1 FROM vpn_brands WHERE vpn_id=?", (vid,)).fetchone()
        if not exists:
            conn.close()
            return await message.answer("❌ VPN Brand not found. Add it first using `/addvpn`.")
        conn.execute("DELETE FROM vpn_packages WHERE vpn_id=? AND pkg_id=?", (vid, pkg))
        conn.execute("INSERT INTO vpn_packages (vpn_id, pkg_id, price) VALUES (?, ?, ?)", (vid, pkg, price))
        conn.commit(); conn.close()
        await message.answer(f"✅ Set VPN Package: {vid} | {pkg} -> {price}৳")
    except: await message.answer("❌ Format Error!")


# --- USER HANDLERS ---

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    await state.clear()
    
    # MAINTENANCE MODE CHECK
    if is_maintenance_mode() and not is_admin(message.from_user.id):
        return await message.answer(get_maintenance_msg(), parse_mode="Markdown")

    # USER OPEN/CLOSE CHECK
    conn = _dbc()
    exists = conn.execute("SELECT 1 FROM users WHERE user_id=?", (message.from_user.id,)).fetchone()
    conn.close()
    
    if not exists and get_bot_status() == 'closed':
        return await message.answer("⚠️ **Notice:**\nCurrently, the bot is closed for new members. Please try again later.")

    get_user_data(message.from_user.id, message.from_user.username, message.from_user.first_name)
    await show_dashboard_ui(message.from_user.id, message.from_user.first_name, bot, message.chat.id)

@dp.callback_query(F.data == "back_home")
async def back_home(c: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await c.answer()
    
    # MAINTENANCE MODE CHECK
    if is_maintenance_mode() and not is_admin(c.from_user.id):
        return await c.message.answer(get_maintenance_msg(), parse_mode="Markdown")

    await c.message.delete()
    await show_dashboard_ui(c.from_user.id, c.from_user.first_name, bot, c.message.chat.id)

@dp.callback_query(F.data == "catalog")
async def show_cat(c: types.CallbackQuery):
    await c.answer()
    if not is_service_enabled("buy_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "🛒 **আইডি ক্রয় সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।",
            parse_mode="Markdown"
        )
    conn = _dbc()
    f6 = conn.execute("SELECT COUNT(*) FROM stock WHERE category='fb61'").fetchone()[0]
    f1 = conn.execute("SELECT COUNT(*) FROM stock WHERE category='fb1000'").fetchone()[0]
    t_id = conn.execute("SELECT COUNT(*) FROM stock WHERE category='tempid'").fetchone()[0]
    
    p6 = get_price('fb61')
    p1 = get_price('fb1000')
    pt = get_price('tempid')
    conn.close()
    
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text=f"🆔 FB 61 ({f6}) ➜ {p6}৳", callback_data="buy_fb61"))
    kb.row(types.InlineKeyboardButton(text=f"🆔 FB 1000 ({f1}) ➜ {p1}৳", callback_data="buy_fb1000"))
    kb.row(types.InlineKeyboardButton(text=f"🆔 Temp ID ({t_id}) ➜ {pt}৳", callback_data="buy_tempid"))
    kb.row(types.InlineKeyboardButton(text="🔙 ফিরে যান", callback_data="back_home"))
    await c.message.edit_text("📥 **আইডি ক্যাটাগরি মেনু**", reply_markup=kb.as_markup())

@dp.callback_query(F.data == "bm_catalog")
async def show_bm_cat(c: types.CallbackQuery):
    await c.answer()
    if not is_service_enabled("buy_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "🛒 **আইডি/বিএম ক্রয় সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।",
            parse_mode="Markdown"
        )
    conn = _dbc()
    ig = conn.execute("SELECT COUNT(*) FROM stock WHERE category='bmig'").fetchone()[0]
    fb = conn.execute("SELECT COUNT(*) FROM stock WHERE category='bmfb'").fetchone()[0]
    pig = get_price('bmig')
    pfb = get_price('bmfb')
    conn.close()
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text=f"📸 Instagram BM ({ig}) ➜ {pig}৳", callback_data="buy_bmig"))
    kb.row(types.InlineKeyboardButton(text=f"📘 Facebook BM ({fb}) ➜ {pfb}৳", callback_data="buy_bmfb"))
    kb.row(types.InlineKeyboardButton(text="🔙 ফিরে যান", callback_data="back_home"))
    await c.message.edit_text("💼 **BM ক্যাটাগরি মেনু**", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("buy_"))
async def buy_qty_start(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    if not is_service_enabled("buy_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "🛒 **আইডি ক্রয় সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।",
            parse_mode="Markdown"
        )
    cat = c.data.split("_")[1]
    await state.update_data(cat=cat)
    await c.message.answer(f"🔢 আপনি কতটি নিতে চান? সংখ্যাটি লিখুন (যেমন: 1):")
    await state.set_state(ShopStates.waiting_for_qty)

@dp.message(ShopStates.waiting_for_qty)
async def process_buy(m: types.Message, state: FSMContext):
    if m.text.startswith("/"): return 
    try:
        raw_qty = to_english_num(m.text.strip())
        if not raw_qty.isdigit(): return await m.answer("⚠️ শুধু সংখ্যা লিখুন।")
        qty = int(raw_qty)
        data = await state.get_data()
        if 'cat' not in data: return await m.answer("❌ Timeout. Try again.")
        cat = data['cat']
        price = get_price(cat)
        total = qty * price
        bal, _, banned = get_user_data(m.from_user.id, m.from_user.username, m.from_user.first_name)

        if banned: return await m.answer("🚫 BANNED.")

        if bal < total:
            _need = max(0, total - bal)
            _kb = InlineKeyboardBuilder()
            _kb.row(types.InlineKeyboardButton(text="⚡ Auto Deposit (bKash/Nagad)", callback_data="dep_auto"))
            _kb.row(types.InlineKeyboardButton(text="🪙 Binance", callback_data="dep_binance"))
            await state.clear()
            await m.answer(
                f"❌ **অপর্যাপ্ত ব্যালেন্স!**\n"
                f"প্রয়োজন: **total৳**\nআছে: **{{bal}}৳**\n"
                f"আরও দরকার: **{{_need}}৳**\n\n"
                f"👇 নিচের button থেকে deposit করুন:",
                reply_markup=_kb.as_markup()
            )
            return

        conn = _dbc(); cursor = conn.cursor()
        items = cursor.execute("SELECT id, data FROM stock WHERE category=? LIMIT ?", (cat, qty)).fetchall()
        if len(items) < qty: conn.close(); return await m.answer(f"❌ স্টক নেই! আছে: {len(items)}টি")

        # [BUYLIMIT_HOOK] 10 pcs / 10 min (FB 1000xx only)
        _bl_ok, _bl_used, _bl_left, _bl_allowed = _bl_allow(m.from_user.id, qty, cat)
        if not _bl_ok:
            try:
                await m.answer(_bl_block_text(_bl_used, _bl_left, qty, _bl_allowed), parse_mode="Markdown")
            except Exception:
                await m.answer(_bl_block_text(_bl_used, _bl_left, qty, _bl_allowed))
            try:
                await state.clear()
            except Exception:
                pass
            return
        conn.execute("UPDATE users SET balance = balance - ? WHERE user_id = ?", (total, m.from_user.id))
        # [BUYLIMIT_HOOK] commit
        try:
            _bl_u, _bl_l = _bl_commit(m.from_user.id, qty, cat)
            if _bl_l:
                await m.answer(_bl_ok_text(_bl_u, _bl_l), parse_mode="Markdown")
        except Exception as _e:
            print(f"[buylimit] commit skip: {_e}")
        
        current_time = datetime.now(timezone(timedelta(hours=6))).strftime("%I:%M %p")
        conn.execute("INSERT INTO sales (user_id, username, category, qty, total, date, time) VALUES (?, ?, ?, ?, ?, ?, ?)",
                       (m.from_user.id, m.from_user.first_name, cat, qty, total, datetime.now().strftime("%Y-%m-%d"), current_time))

        _lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID","ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}.get(cat, cat.upper())
        # [DELIVERY_FORMAT_PATCH_V1] — ask format before dumping
        # Get sale_id (last inserted), delete stock, archive, then ask format
        _sale_id = cursor.execute("SELECT last_insert_rowid()").fetchone()[0]
        _now_ts = int(__import__("time").time())
        _uname = (f"@{m.from_user.username}" if m.from_user.username else None)
        _delivered = []
        for i in items:
            cursor.execute("DELETE FROM stock WHERE id = ?", (i[0],))
            _delivered.append((i[0], i[1]))
            try:
                conn.execute(
                    "INSERT INTO delivery_archive (sale_id, user_id, username, category, stock_id, data, source, delivered_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (_sale_id, m.from_user.id, _uname, cat, i[0], i[1], 'bot', _now_ts),
                )
            except Exception:
                pass
        conn.commit()
        conn.close()

        # Stash for callback
        try:
            _PENDING_DELIVERY[_sale_id] = {
                "user_id": m.from_user.id,
                "cat": cat, "lbl": _lbl, "qty": qty,
                "items": _delivered,
                "ts": _now_ts,
            }
        except NameError:
            pass

        if cat == "tempid":
            await m.answer(
                "⚠️ **Temp ID — গুরুত্বপূর্ণ নিয়ম**\n\n"
                "⏱ Replace time: **2 ঘণ্টা**\n\n"
                "❌ Verify হয়ে গেলে replace **হবে না**\n"
                "✅ শুধু **login issue** হলে replace সম্ভব\n\n"
                "নিয়মের বাইরে replace request দিলে **reject** করা হবে।"
            )

        _kb = types.InlineKeyboardMarkup(inline_keyboard=[
            [
                types.InlineKeyboardButton(text="📊 Excel (.xlsx)", callback_data=f"dfmt:xlsx:{_sale_id}"),
                types.InlineKeyboardButton(text="📝 Text (.txt)",   callback_data=f"dfmt:txt:{_sale_id}"),
            ],
            [
                types.InlineKeyboardButton(text="🔙 মূল মেনু (Home)", callback_data="back_home"),
            ]
        ])
        report_time = "2 Hours" if qty < 10 else "6 Hours"

        # Instant text delivery directly in chat for up to 3 accounts
        _instant_block = ""
        if qty <= 3 and _delivered:
            _items_txt = []
            for _idx, _item in enumerate(_delivered, 1):
                _raw = _item[1].strip() if _item else ""
                _items_txt.append(f"🔹 **আইডি #{_idx}:**\n`{_raw}`")
            _instant_block = "\n\n📦 **আপনার ইনস্ট্যান্ট ডেলিভারি:**\n" + "\n\n".join(_items_txt) + "\n"

        await m.answer(
            f"✅ **পেমেন্ট সফল!** — {_lbl} × {qty}\n"
            f"⏱ রিপোর্ট টাইম: {report_time} • 🔐 লগইন গ্যারান্টি"
            f"{_instant_block}\n"
            f"📥 ফাইল ডাউনলোড করতে চাইলে নিচের ফরম্যাট বেছে নিন:",
            reply_markup=_kb,
            parse_mode="Markdown"
        )
        await state.clear()
    except Exception as e:
        await state.clear(); await m.answer(f"❌ ত্রুটি: {e}")

@dp.callback_query(F.data == "deposit")
async def dep_start(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    if not is_service_enabled("deposit_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "💰 **ডিপোজিট সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।",
            parse_mode="Markdown"
        )
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="⚡ Auto Payment (bKash / Nagad)", callback_data="dep_auto"))
    kb.row(types.InlineKeyboardButton(text="💎 Binance USDT", callback_data="dep_binance"))
    await c.message.answer(
        "🌟 *ডিপোজিট পদ্ধতি বাছাই করুন* 🌟\n\n"
        "━━━━━━━━━━━━━━━━━━━━━\n"
        "⚡ *Auto Payment* — সবচেয়ে দ্রুত ও সহজ\n"
        "   ▸ bKash / Nagad সাপোর্ট\n"
        "   ▸ এক ক্লিকে পেমেন্ট পেজ\n"
        "   ▸ ১০-৩০ সেকেন্ডে ব্যালেন্স যোগ\n"
        "   ▸ মিনিমাম ১০৳\n\n"
        "💎 *Binance USDT* — ডলারে পেমেন্ট\n"
        "   ▸ Rate: *$1 = 125৳* (fixed)\n"
        "   ▸ Screenshot পাঠাবেন\n"
        "   ▸ Admin ম্যানুয়ালি অ্যাপ্রুভ করবে\n"
        "   ▸ মিনিমাম 1$\n"
        "━━━━━━━━━━━━━━━━━━━━━",
        reply_markup=kb.as_markup(),
        parse_mode="Markdown"
    )

@dp.callback_query(F.data == "dep_auto")
async def dep_auto_start(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    if not is_service_enabled("deposit_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "💰 **ডিপোজিট সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।",
            parse_mode="Markdown"
        )
    await c.message.answer(
        "⚡ *Auto Payment*\n\n"
        "💰 কত টাকা ডিপোজিট করবেন? \n"
        "শুধু সংখ্যা লিখুন (যেমন: `100`)\n\n"
        "ℹ️ মিনিমাম ১০৳ — Submit করলেই পেমেন্ট লিংক পাবেন।",
        parse_mode="Markdown"
    )
    await state.update_data(deposit_method="auto")
    await state.set_state(ShopStates.waiting_for_deposit_amount)

@dp.callback_query(F.data == "dep_binance")
async def dep_binance_start(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    if not is_service_enabled("deposit_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "💰 **ডিপোজিট সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।",
            parse_mode="Markdown"
        )
    await c.message.answer(
        "💎 *Binance USDT Deposit*\n\n"
        "━━━━━━━━━━━━━━━━━━━━━\n"
        f"🏦 *Binance UID:* `{BINANCE_ID}`\n"
        "💱 *Rate:* `$1 = 125৳` (fixed)\n"
        "━━━━━━━━━━━━━━━━━━━━━\n\n"
        "💵 কত ডলার পাঠাবেন? \n"
        "শুধু সংখ্যা লিখুন (যেমন: `5` মানে $5)\n\n"
        "ℹ️ মিনিমাম 1$ = 125৳",
        parse_mode="Markdown"
    )
    await state.update_data(deposit_method="binance")
    await state.set_state(ShopStates.waiting_for_deposit_amount)

@dp.message(ShopStates.waiting_for_deposit_amount)
async def dep_amt(m: types.Message, state: FSMContext):
    if m.text.startswith("/"): return
    val_str = to_english_num(m.text).strip()
    _data = await state.get_data()
    _method = _data.get("deposit_method", "auto")

    # ============ AUTO (ZiniPay) ============
    if _method == "auto":
        if not val_str.isdigit() or int(val_str) < 10:
            return await m.answer("⚠️ মিনিমাম *১০ টাকা*। শুধু সংখ্যা লিখুন (যেমন: `100`)", parse_mode="Markdown")
        amt = int(val_str)
        username = f"@{m.from_user.username}" if m.from_user.username else m.from_user.first_name
        wait_msg = await m.answer("⏳ পেমেন্ট লিংক তৈরি হচ্ছে...")
        payment_url = create_zinipay_invoice(m.from_user.id, username, amt)
        try: await wait_msg.delete()
        except: pass
        if not payment_url:
            await m.answer("❌ পেমেন্ট gateway এ সমস্যা। একটু পরে আবার চেষ্টা করুন অথবা 💎 Binance ব্যবহার করুন।")
            await state.clear()
            return
        kb = InlineKeyboardBuilder()
        kb.row(types.InlineKeyboardButton(text=f"💳 এখনই পেমেন্ট করুন — {amt}৳", url=payment_url))
        await m.answer(
            f"✨ *পেমেন্ট লিংক রেডি* ✨\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n"
            f"💰 Amount: *{amt}৳*\n"
            f"⚡ Method: bKash / Nagad (Auto)\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"🔗 নিচের বোতামে ট্যাপ করুন — সরাসরি পেমেন্ট পেজে চলে যাবেন।\n\n"
            f"✅ পেমেন্ট সফল হলে *১০-৩০ সেকেন্ডে* ব্যালেন্স অটো যোগ হবে।\n"
            f"🔔 কোনো screenshot বা confirmation message পাঠাতে হবে না।",
            reply_markup=kb.as_markup(),
            parse_mode="Markdown",
            disable_web_page_preview=True
        )
        await state.clear()
        return

    # ============ BINANCE ($→BDT→screenshot) ============
    if _method == "binance":
        # ডলার amount পাছ — শুধু সংখ্যা / দশমিক
        clean = val_str.replace("$", "").replace("usd", "").replace("USD", "").strip()
        try:
            usd = float(clean)
        except:
            return await m.answer("⚠️ শুধু সংখ্যা লিখুন (যেমন: `5` মানে $5)", parse_mode="Markdown")
        if usd < 1:
            return await m.answer("⚠️ মিনিমাম *$1*। আবার লিখুন।", parse_mode="Markdown")
        bdt_amount = int(round(usd * 125))
        await state.update_data(amount_text=str(bdt_amount), usd_amount=str(usd))
        await m.answer(
            f"💎 *Binance Deposit Confirm*\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n"
            f"💵 You send: *${usd}*\n"
            f"💱 Rate: `$1 = 125৳`\n"
            f"💰 You'll get: *{bdt_amount}৳*\n"
            f"━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"🏦 Binance UID: `{BINANCE_ID}`\n\n"
            f"📸 *এখন পেমেন্টের screenshot পাঠান* (বাধ্যতামূলক)\n"
            f"⚠️ Screenshot ছাড়া রিকোয়েস্ট গ্রহণ হবে না।",
            parse_mode="Markdown"
        )
        await state.set_state(ShopStates.waiting_for_screenshot)
        return

    # Fallback
    await m.answer("⚠️ Session expired। আবার /start দিন।")
    await state.clear()

@dp.message(ShopStates.waiting_for_screenshot)
async def dep_submit(m: types.Message, state: FSMContext):
    if m.text and m.text.startswith("/"): return

    if not m.photo:
        return await m.answer("❌ আপনি সঠিক নিয়ম মানেননি। পেমেন্ট স্ক্রিনশট দেওয়া বাধ্যতামূলক।")

    data = await state.get_data()
    now_ts = datetime.now(timezone.utc).timestamp()
    
    clean_amt = ''.join(filter(lambda x: x.isdigit(), data.get('amount_text', '0')))
    amt = int(clean_amt) if clean_amt else 0
    
    # --- DEPOSIT ANTI-SPAM LOGIC ---
    conn = _dbc()
    last_dep = conn.execute("SELECT timestamp, amount FROM payment_logs WHERE user_id=? ORDER BY timestamp DESC LIMIT 1", (m.from_user.id,)).fetchone()
    
    if last_dep and last_dep[0]:
        time_diff = now_ts - last_dep[0]
        if time_diff < 300: 
            conn.close()
            if last_dep[1] == amt:
                return await m.answer("⚠️ **Warning:** You are submitting the same deposit request too quickly! Please wait 5 minutes before submitting again to avoid double transactions.")
            else:
                return await m.answer("⏳ Please wait at least 5 minutes between deposit requests.")

    req_id = str(uuid.uuid4())[:8]
    username_display = f"@{m.from_user.username}" if m.from_user.username else "No Username"
    real_name = m.from_user.first_name
    admins = conn.execute("SELECT user_id FROM admins").fetchall()

    try:
        conn.execute("INSERT INTO payment_logs (req_id, user_id, username, amount, status, date, admin_name, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                     (req_id, m.from_user.id, username_display, amt, 'pending', datetime.now().strftime("%Y-%m-%d"), "None", now_ts))
        conn.commit()
    except Exception as e:
        await m.answer(f"❌ Error Saving Request: {e}")
        conn.close()
        return
    conn.close()

    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text=f"✅ Add {amt}৳", callback_data=f"pay_ok_{req_id}_{amt}"),
           types.InlineKeyboardButton(text="❌ Reject", callback_data=f"pay_no_{req_id}"))

    admin_msg = f"🔔 **Pay Request**\n👤 {real_name} ({username_display})\n🆔 `{m.from_user.id}`\n💰 {amt}৳\n📱 `{data.get('sender')}`"

    for admin in admins:
        try: await bot.send_photo(admin[0], m.photo[-1].file_id, caption=admin_msg, reply_markup=kb.as_markup())
        except: pass

    await m.answer("⏳ জমা হয়েছে! এডমিন চেক করে ব্যালেন্স দিবে।")
    await state.clear()
    await show_dashboard_ui(m.from_user.id, m.from_user.first_name, bot, m.chat.id)

@dp.callback_query(F.data.startswith("pay_"))
async def admin_pay_action(c: types.CallbackQuery):
    await c.answer()
    
    try: await c.message.edit_reply_markup(reply_markup=None)
    except: pass
    
    parts = c.data.split("_")
    action = parts[1] 
    req_id = parts[2]

    conn = _dbc()
    req_data = conn.execute("SELECT user_id, status, admin_name FROM payment_logs WHERE req_id=?", (req_id,)).fetchone()

    if not req_data:
        conn.close()
        return await c.message.answer("❌ Error: Request not found.")

    user_id, status, admin_name = req_data

    if status != 'pending':
        conn.close()
        try: await c.message.edit_caption(caption=f"{c.message.caption}\n\n⚠️ **Already {status.upper()} by {admin_name}!**")
        except: pass
        return await c.message.reply(f"⚠️ **Already {status.upper()} by {admin_name}!**")

    current_admin_name = c.from_user.first_name

    if action == "ok":
        amount = int(parts[3])
        conn.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (amount, user_id))
        conn.execute("UPDATE payment_logs SET status='approved', amount=?, admin_name=? WHERE req_id=?", (amount, current_admin_name, req_id))
        conn.commit()
        
        await bot.send_message(user_id, f"✅ **Received!** Balance Added: {amount}৳")
        try: await c.message.edit_caption(caption=f"{c.message.caption}\n\n✅ **Approved by {current_admin_name}**")
        except: pass
    else:
        conn.execute("UPDATE payment_logs SET status='rejected', admin_name=? WHERE req_id=?", (current_admin_name, req_id))
        conn.commit()
        
        await bot.send_message(user_id, "❌ **Rejected!** Invalid Transaction.")
        try: await c.message.edit_caption(caption=f"{c.message.caption}\n\n❌ **Rejected by {current_admin_name}**")
        except: pass

    conn.close()

# --- VPN FEATURES ---

@dp.callback_query(F.data == "vpn_catalog")
async def show_vpn_catalog(c: types.CallbackQuery):
    await c.answer()
    if not is_service_enabled("vpn_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "🛡️ **ভিপিএন সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।",
            parse_mode="Markdown"
        )
    
    conn = _dbc()
    brands = conn.execute("SELECT DISTINCT v.vpn_id, v.vpn_name FROM vpn_brands v JOIN vpn_packages p ON v.vpn_id = p.vpn_id").fetchall()
    conn.close()
    
    kb = InlineKeyboardBuilder()
    
    for i in range(0, len(brands), 2):
        row_buttons = []
        for j in range(2):
            if i + j < len(brands):
                vpn_id, vpn_name = brands[i+j]
                emoji = VPN_EMOJIS.get(vpn_id, "⚛️")
                row_buttons.append(types.InlineKeyboardButton(text=f"{emoji} {vpn_name}", callback_data=f"vpn_sel_{vpn_id}"))
        kb.row(*row_buttons)
        
    kb.row(types.InlineKeyboardButton(text="❌ Close", callback_data="back_home"))
    
    msg = "🌐 **ভিপিএন ব্র্যান্ড সিলেক্ট করুন:**\n━━━━━━━━━━━━━━━━━━━━\n👇 আপনার পছন্দের ভিপিএন এর উপর ক্লিক করুন:"
    await c.message.edit_text(msg, reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("vpn_sel_"))
async def show_vpn_packages(c: types.CallbackQuery):
    await c.answer()
    vpn_id = c.data.split("_")[2]
    
    conn = _dbc()
    brand = conn.execute("SELECT vpn_name FROM vpn_brands WHERE vpn_id=?", (vpn_id,)).fetchone()
    pkgs = conn.execute("SELECT pkg_id, price FROM vpn_packages WHERE vpn_id=?", (vpn_id,)).fetchall()
    conn.close()
    
    if not brand: return await c.message.edit_text("❌ VPN not found.")
    brand_name = brand[0]
    emoji = VPN_EMOJIS.get(vpn_id, "⚛️")
    
    kb = InlineKeyboardBuilder()
    for pkg_id, price in pkgs:
        pkg_name = format_pkg_name(pkg_id)
        icon = "🟢" if "7" in pkg_id else "🟡" if "9" in pkg_id else "🔵" if "14" in pkg_id else "🟣"
        kb.row(types.InlineKeyboardButton(text=f"{icon} {pkg_name} 📦 | {price}৳", callback_data=f"vpn_pkg_{vpn_id}_{pkg_id}"))
        
    kb.row(types.InlineKeyboardButton(text="🔙 Back", callback_data="vpn_catalog"),
           types.InlineKeyboardButton(text="❌ Close", callback_data="back_home"))
    
    msg = f"{emoji} **{brand_name} Premium Plans:**\n━━━━━━━━━━━━━━━━━━━━\n👇 আপনার প্রয়োজনীয় প্যাকেজটি নির্বাচন করুন:"
    await c.message.edit_text(msg, reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("vpn_pkg_"))
async def vpn_confirm_order(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    parts = c.data.split("_")
    vpn_id = parts[2]
    pkg_id = parts[3]
    
    conn = _dbc()
    brand = conn.execute("SELECT vpn_name FROM vpn_brands WHERE vpn_id=?", (vpn_id,)).fetchone()
    pkg = conn.execute("SELECT price FROM vpn_packages WHERE vpn_id=? AND pkg_id=?", (vpn_id, pkg_id)).fetchone()
    conn.close()
    
    if not brand or not pkg: return await c.message.edit_text("❌ Error loading package.")
    
    vpn_name = brand[0]
    price = pkg[0]
    pkg_name = format_pkg_name(pkg_id)
    emoji = VPN_EMOJIS.get(vpn_id, "⚛️")
    
    bal, _, banned = get_user_data(c.from_user.id, c.from_user.username, c.from_user.first_name)
    
    if banned: return await c.message.answer("🚫 **YOU ARE BANNED** 🚫")
        
    if bal < price:
        _need = max(0, price - bal)
        _kb = InlineKeyboardBuilder()
        _kb.row(types.InlineKeyboardButton(text="⚡ Auto Deposit (bKash/Nagad)", callback_data="dep_auto"))
        _kb.row(types.InlineKeyboardButton(text="🪙 Binance", callback_data="dep_binance"))
        await state.clear()
        await c.message.answer(
            f"❌ **অপর্যাপ্ত ব্যালেন্স!**\n"
            f"প্রয়োজন: {price}৳\n"
            f"আছে: {bal}৳\n"
            f"আরও লাগবে: {_need}৳\n\n"
            f"⚠️ **মিনিমাম ডিপোজিট ১০৳ (Binance 1$)**\n\n"
            f"👇 নিচের অপশন থেকে ডিপোজিট করুন:",
            reply_markup=_kb.as_markup()
        )
        return
        
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="✅ Confirm", callback_data=f"vpn_buy_{vpn_id}_{pkg_id}"))
    kb.row(types.InlineKeyboardButton(text="🔙 Back", callback_data=f"vpn_sel_{vpn_id}"),
           types.InlineKeyboardButton(text="❌ Close", callback_data="back_home"))
           
    msg = f"┏━━━━━━━━━━━━━━━━━━━━━┓\n┣ {emoji} **{vpn_name}**\n┗━━━━━━━━━━━━━━━━━━━━━┛\n📦 **প্যাকেজ:** {pkg_name}\n💰 **মূল্য:** {price} টাকা\n\n⚠️ **অর্ডার কনফার্ম করবেন?**"
    await c.message.edit_text(msg, reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("vpn_buy_"))
async def process_vpn_buy(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    
    try: await c.message.edit_reply_markup(reply_markup=None)
    except: pass
    
    parts = c.data.split("_")
    vpn_id = parts[2]
    pkg_id = parts[3]
    
    conn = _dbc()
    brand = conn.execute("SELECT vpn_name FROM vpn_brands WHERE vpn_id=?", (vpn_id,)).fetchone()
    pkg = conn.execute("SELECT price FROM vpn_packages WHERE vpn_id=? AND pkg_id=?", (vpn_id, pkg_id)).fetchone()
    
    if not brand or not pkg: 
        conn.close()
        return await c.message.answer("❌ Error processing package.")
        
    vpn_name = brand[0]
    price = pkg[0]
    pkg_name = format_pkg_name(pkg_id)
    emoji = VPN_EMOJIS.get(vpn_id, "⚛️")
    
    bal, _, banned = get_user_data(c.from_user.id, c.from_user.username, c.from_user.first_name)
    if banned: 
        conn.close()
        return
    
    if bal < price:
        conn.close()
        return await c.message.answer("❌ আপনার একাউন্টে পর্যাপ্ত ব্যালেন্স নেই।")




    # === NORD_AUTO_DELIVER_V5 ===
    # V5: global VPN service ON/OFF (config: vpn_service_enabled) + Nord auto-deliver.
    import json as _json
    # --- Global VPN service gate (covers ALL vpn brands: Nord, Express, etc.) ---
    try:
        _svc_row = conn.execute(
            "SELECT value FROM config WHERE key='vpn_service_enabled'"
        ).fetchone()
    except Exception:
        _svc_row = None
    _svc_val = (str(_svc_row[0]).strip().lower() if _svc_row else 'on')
    if _svc_val in ('0', 'off', 'false', 'no', 'closed', 'disabled'):
        conn.close()
        return await c.message.answer(
            "🛡️ VPN সার্ভিস এই মুহূর্তে সাময়িকভাবে বন্ধ আছে।\n"
            "⏳ শীঘ্রই আবার চালু হবে — একটু অপেক্ষা করুন।\n"
            "💙 ধন্যবাদ আপনার ধৈর্যের জন্য!"
        )
    if vpn_id == 'nord':
        try:
            _nord_row = conn.execute(
                """SELECT id, data FROM nord_stock
                   WHERE pkg_id = ? AND delivered_count < 2
                     AND id NOT IN (SELECT stock_id FROM nord_deliveries WHERE user_id = ?)
                   ORDER BY delivered_count ASC, id ASC LIMIT 1""",
                (pkg_id, c.from_user.id),
            ).fetchone()
        except Exception:
            _nord_row = None
        if _nord_row:
            _stock_id, _vpn_info = _nord_row
            _order_id = str(uuid.uuid4())[:8]
            _now_ts = int(datetime.now().timestamp())
            _cur_time = datetime.now(timezone(timedelta(hours=6))).strftime("%I:%M %p")
            _uname = f"@{c.from_user.username}" if c.from_user.username else f"User {c.from_user.id}"
            # Dynamic package icon based on duration
            _pkg_lower = (str(pkg_name) + " " + str(pkg_id)).lower()
            if '7' in _pkg_lower and 'day' in _pkg_lower:
                _pkg_icon = "⏱️"
            elif '30' in _pkg_lower or 'month' in _pkg_lower:
                _pkg_icon = "📅"
            elif 'year' in _pkg_lower or '365' in _pkg_lower:
                _pkg_icon = "🗓️"
            else:
                _pkg_icon = "📦"
            try:
                conn.execute("BEGIN IMMEDIATE")
                _upd = conn.execute(
                    "UPDATE users SET balance = balance - ? WHERE user_id = ? AND balance >= ?",
                    (price, c.from_user.id, price),
                )
                if _upd.rowcount != 1:
                    raise Exception("Insufficient balance during auto-deliver")
                conn.execute(
                    "UPDATE nord_stock SET delivered_count = delivered_count + 1 WHERE id = ? AND delivered_count < 2",
                    (_stock_id,),
                )
                conn.execute(
                    "INSERT INTO nord_deliveries (stock_id, user_id, order_id, delivered_at) VALUES (?, ?, ?, ?)",
                    (_stock_id, c.from_user.id, _order_id, _now_ts),
                )
                conn.execute(
                    "INSERT INTO sales (user_id, username, category, qty, total, date, time) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (c.from_user.id, _uname, f"VPN: {vpn_name}", 1, price,
                     datetime.now().strftime("%Y-%m-%d"), _cur_time),
                )
                conn.execute(
                    "INSERT INTO vpn_orders (order_id, user_id, vpn_name, duration, price, status, date, admin_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (_order_id, c.from_user.id, vpn_name, pkg_name, price,
                     'delivered', datetime.now().strftime("%Y-%m-%d"), 'AUTO'),
                )
                try:
                    conn.execute(
                        "INSERT INTO delivery_archive (sale_id, user_id, username, category, stock_id, data, source, delivered_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                        (None, c.from_user.id,
                         (f"@{c.from_user.username}" if c.from_user.username else None),
                         f"VPN: {vpn_name}", _stock_id, _vpn_info, 'bot-auto', _now_ts),
                    )
                except Exception:
                    pass
                conn.commit()
            except Exception as _e:
                try: conn.rollback()
                except: pass
                conn.close()
                return await c.message.answer(f"❌ Auto-delivery ব্যর্থ: {_e}\nAdmin কে জানান।")

            # Parse structured data (V3+) or fallback to raw text (legacy rows)
            _email = ""
            _password = ""
            try:
                _d = _json.loads(_vpn_info)
                _email = str(_d.get("email", "")).strip()
                _password = str(_d.get("password", "")).strip()
            except Exception:
                _email = ""
                _password = ""
            if _email and _password:
                _user_msg = (
                    f"🛡️ NordVPN Premium Account 🛡️\n\n"
                    f"📧 Email: `{_email}`\n"
                    f"🔑 Password: `{_password}`\n\n"
                    f"{_pkg_icon} Package: {pkg_name}\n"
                    f"⚠️ পাসওয়ার্ড বা কোনো তথ্য পরিবর্তন করবেন না।\n"
                    f"🛠️ যেকোনো সমস্যায় দ্রুত নক দিন, সাপোর্ট পাবেন।\n"
                    f"💙 ধন্যবাদ!\n\n"
                    f"🆔 Order: `{_order_id}`"
                )
            else:
                _user_msg = (
                    f"🎉 **আপনার VPN অর্ডার ডেলিভারি সম্পন্ন হয়েছে!** 🎉\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"🛡️ **ব্র্যান্ড:** {vpn_name}\n"
                    f"{_pkg_icon} **প্যাকেজ:** {pkg_name}\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"🔐 **আপনার লগইন ডিটেইলস:**\n"
                    f"```text\n{_vpn_info}\n```\n"
                    f"🆔 Order: `{_order_id}`"
                )
            try:
                await c.message.answer(_user_msg, parse_mode="Markdown")
            except Exception:
                await c.message.answer(_user_msg)

            try:
                _admins = conn.execute("SELECT user_id FROM admins").fetchall()
            except Exception:
                _admins = []
            conn.close()
            _udisp = f"@{c.from_user.username}" if c.from_user.username else "No Username"
            _admin_notify = (
                f"⚡ **AUTO-DELIVERED (🛡️ NordVPN)**\n"
                f"👤 {c.from_user.first_name} | {_udisp} | `{c.from_user.id}`\n"
                f"{_pkg_icon} {pkg_name} | 💰 {price}৳\n"
                f"🆔 Order: `{_order_id}` | Stock #{_stock_id}"
            )
            for _a in _admins:
                try: await bot.send_message(_a[0], _admin_notify)
                except: pass

            # --- Low-stock warning (rate-limited to 1 alert per pkg per hour) ---
            try:
                _wconn = _dbc()
                _remaining = _wconn.execute(
                    "SELECT COALESCE(SUM(2 - delivered_count), 0) FROM nord_stock "
                    "WHERE pkg_id = ? AND delivered_count < 2",
                    (pkg_id,),
                ).fetchone()[0]
                _thr_row = _wconn.execute(
                    "SELECT value FROM config WHERE key = 'nord_warn_threshold'"
                ).fetchone()
                _thr = int(_thr_row[0]) if _thr_row else 3
                if _remaining <= _thr:
                    _last_row = _wconn.execute(
                        "SELECT value FROM config WHERE key = ?",
                        (f"nord_last_alert_{pkg_id}",),
                    ).fetchone()
                    _last = int(_last_row[0]) if _last_row else 0
                    if _now_ts - _last >= 3600:
                        _wconn.execute(
                            "INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)",
                            (f"nord_last_alert_{pkg_id}", str(_now_ts)),
                        )
                        _wconn.commit()
                        _wconn.close()
                        _lvl = "🚨 OUT OF STOCK" if _remaining <= 0 else "⚠️ Stock LOW"
                        _warn = (
                            f"{_lvl} — 🛡️ NordVPN\n"
                            f"{_pkg_icon} Package: {pkg_name} ({pkg_id})\n"
                            f"🔻 Remaining slots: {_remaining}\n"
                            f"➕ Panel: /nord এ গিয়ে stock refill করুন।"
                        )
                        for _a in _admins:
                            try: await bot.send_message(_a[0], _warn)
                            except: pass
                    else:
                        _wconn.close()
                else:
                    _wconn.close()
            except Exception:
                try: _wconn.close()
                except: pass
            return
    # === END NORD_AUTO_DELIVER_V5 ===

    conn.execute("UPDATE users SET balance = balance - ? WHERE user_id = ?", (price, c.from_user.id))
        
    order_id = str(uuid.uuid4())[:8]
    username_display = f"@{c.from_user.username}" if c.from_user.username else "No Username"
    real_name = c.from_user.first_name
    
    try:
        conn.execute("INSERT INTO vpn_orders (order_id, user_id, vpn_name, duration, price, status, date, admin_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                     (order_id, c.from_user.id, vpn_name, pkg_name, price, 'pending', datetime.now().strftime("%Y-%m-%d"), "None"))
        conn.commit()
    except Exception as e:
        await c.message.answer(f"❌ Error: {e}")
        conn.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (price, c.from_user.id))
        conn.commit()
        conn.close()
        return

    admins = conn.execute("SELECT user_id FROM admins").fetchall()
    conn.close()
    
    admin_msg = (
        f"┏━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┣ 🚨 **NEW VPN ORDER** 🚨\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━┛\n"
        f"👤 **Name:** {real_name}\n"
        f"🔗 **User:** {username_display}\n"
        f"🆔 **ID:** `{c.from_user.id}`\n\n"
        f"{emoji} **Brand:** {vpn_name}\n"
        f"📦 **Package:** {pkg_name}\n"
        f"💰 **Price:** {price} BDT\n"
        f"━━━━━━━━━━━━━━━━━━━━━"
    )
    
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="📦 Deliver", callback_data=f"vpn_deliv_{order_id}"),
           types.InlineKeyboardButton(text="❌ Stock Out", callback_data=f"vpn_out_{order_id}"))
    
    for admin in admins:
        try: await bot.send_message(admin[0], admin_msg, reply_markup=kb.as_markup())
        except: pass
        
    user_msg = f"✅ **অর্ডার সফল!**\nঅ্যাডমিন চেক করে ইনফো সেন্ড করবে।"
    await c.message.answer(user_msg)
    
    asyncio.create_task(vpn_delivery_reminder(order_id, admins, admin_msg))

async def vpn_delivery_reminder(order_id, admins, original_msg):
    await asyncio.sleep(600) 
    conn = _dbc()
    status = conn.execute("SELECT status FROM vpn_orders WHERE order_id=?", (order_id,)).fetchone()
    conn.close()
    
    if status and status[0] == 'pending':
        reminder_msg = f"⏰ **REMINDER (10 Min)! PENDING VPN ORDER:**\n\n{original_msg}"
        kb = InlineKeyboardBuilder()
        kb.row(types.InlineKeyboardButton(text="📦 Deliver", callback_data=f"vpn_deliv_{order_id}"),
               types.InlineKeyboardButton(text="❌ Stock Out", callback_data=f"vpn_out_{order_id}"))
        for admin in admins:
            try: await bot.send_message(admin[0], reminder_msg, reply_markup=kb.as_markup())
            except: pass
            
        await asyncio.sleep(1800) 
        conn = _dbc()
        status = conn.execute("SELECT status FROM vpn_orders WHERE order_id=?", (order_id,)).fetchone()
        conn.close()
        if status and status[0] == 'pending':
            urgent_msg = f"🚨 **URGENT (40 Min)! PENDING VPN ORDER:**\n\n{original_msg}"
            for admin in admins:
                try: await bot.send_message(admin[0], urgent_msg, reply_markup=kb.as_markup())
                except: pass

@dp.callback_query(F.data.startswith("vpn_deliv_"))
async def start_vpn_delivery(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    order_id = c.data.split("_")[2]
    
    conn = _dbc()
    order = conn.execute("SELECT status, admin_name FROM vpn_orders WHERE order_id=?", (order_id,)).fetchone()
    
    if not order: 
        conn.close()
        return await c.message.answer("❌ Order not found.")
    
    if order[0] != 'pending':
        conn.close()
        try: await c.message.edit_reply_markup(reply_markup=None)
        except: pass
        try: await c.message.edit_text(f"{c.message.text}\n\n✅ **Processed by {order[1]}!**")
        except: pass
        return await c.message.reply(f"⚠️ **Already processing/delivered by {order[1]}!**")
        
    conn.execute("UPDATE vpn_orders SET status='processing', admin_name=? WHERE order_id=?", (c.from_user.first_name, order_id))
    conn.commit()
    conn.close()
    
    await state.update_data(current_vpn_order=order_id, deliver_msg_id=c.message.message_id)
    await c.message.reply("✍️ **একাউন্টের ডিটেইলস দিন (Email:Pass):**\nঅথবা স্টক না থাকলে `/stockout` লিখুন।")
    await state.set_state(ShopStates.waiting_for_vpn_delivery)

@dp.callback_query(F.data.startswith("vpn_out_"))
async def vpn_stock_out_btn(c: types.CallbackQuery):
    await c.answer()
    order_id = c.data.split("_")[2]
    
    conn = _dbc()
    order = conn.execute("SELECT user_id, price, status, admin_name, vpn_name, duration FROM vpn_orders WHERE order_id=?", (order_id,)).fetchone()
    
    if not order:
        conn.close()
        return await c.message.answer("❌ Order not found.")
        
    user_id, price, status, admin_name, vpn_name, duration = order
    
    if status != 'pending':
        conn.close()
        try: await c.message.edit_reply_markup(reply_markup=None)
        except: pass
        return await c.message.reply(f"⚠️ **Already processed by {admin_name}!**")
        
    admin_first_name = c.from_user.first_name
    
    conn.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (price, user_id))
    conn.execute("UPDATE vpn_orders SET status='stock_out', admin_name=? WHERE order_id=?", (admin_first_name, order_id))
    conn.commit()
    conn.close()
    
    try: await c.message.edit_reply_markup(reply_markup=None)
    except: pass
    try: await c.message.edit_text(f"{c.message.text}\n\n❌ **Marked OUT OF STOCK by {admin_first_name}**")
    except: pass
    
    user_msg = (
        f"┏━━━━━━━━━━━━━━━━━━━━━┓\n"
        f"┣ ⚠️ **ORDER CANCELLED**\n"
        f"┗━━━━━━━━━━━━━━━━━━━━━┛\n"
        f"😔 **দুঃখিত! স্টক আউট!**\n\n"
        f"আপনার অর্ডারকৃত **{vpn_name} ({duration})** বর্তমানে আমাদের স্টকে নেই।\n"
        f"💰 আপনার অ্যাকাউন্টে **{price}৳** রিফান্ড করা হয়েছে।\n\n"
        f"অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।"
    )
    try: await bot.send_message(user_id, user_msg)
    except: pass

@dp.message(ShopStates.waiting_for_vpn_delivery)
async def submit_vpn_delivery(m: types.Message, state: FSMContext):
    data = await state.get_data()
    order_id = data.get('current_vpn_order')
    msg_id = data.get('deliver_msg_id')
    
    if not order_id: return await m.answer("❌ Error: Order session lost.")
    
    if m.text and (m.text.lower() == "/stockout" or m.text.lower() == "/cancel"):
        conn = _dbc()
        order = conn.execute("SELECT user_id, price, vpn_name, duration FROM vpn_orders WHERE order_id=?", (order_id,)).fetchone()
        if order:
            user_id, price, vpn_name, duration = order
            conn.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (price, user_id))
            conn.execute("UPDATE vpn_orders SET status='stock_out', admin_name=? WHERE order_id=?", (m.from_user.first_name, order_id))
            conn.commit()
            
            user_msg = (
                f"┏━━━━━━━━━━━━━━━━━━━━━┓\n"
                f"┣ ⚠️ **ORDER CANCELLED**\n"
                f"┗━━━━━━━━━━━━━━━━━━━━━┛\n"
                f"😔 **দুঃখিত! স্টক আউট!**\n\n"
                f"আপনার অর্ডারকৃত **{vpn_name} ({duration})** বর্তমানে আমাদের স্টকে নেই।\n"
                f"💰 আপনার অ্যাকাউন্টে **{price}৳** রিফান্ড করা হয়েছে।\n\n"
                f"অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন।"
            )
            try: await bot.send_message(user_id, user_msg)
            except: pass
        conn.close()
        await state.clear()
        try: await bot.edit_message_text(f"❌ **Marked OUT OF STOCK by {m.from_user.first_name}**", chat_id=m.chat.id, message_id=msg_id)
        except: pass
        return await m.answer("✅ Marked as Stock Out and refunded user.")
    
    vpn_info = m.text
    admin_name = m.from_user.first_name
    
    conn = _dbc()
    order = conn.execute("SELECT user_id, price, vpn_name, duration FROM vpn_orders WHERE order_id=?", (order_id,)).fetchone()
    
    if not order:
        conn.close()
        await state.clear()
        return await m.answer("❌ Order not found.")
        
    user_id, price, vpn_name, duration = order
    emoji = next((v for k, v in VPN_EMOJIS.items() if k.lower() in vpn_name.lower()), "⚛️")
    
    current_time = datetime.now(timezone(timedelta(hours=6))).strftime("%I:%M %p")
    conn.execute("INSERT INTO sales (user_id, username, category, qty, total, date, time) VALUES (?, ?, ?, ?, ?, ?, ?)",
                 (user_id, f"User {user_id}", f"VPN: {vpn_name}", 1, price, datetime.now().strftime("%Y-%m-%d"), current_time))
    conn.execute("UPDATE vpn_orders SET status='delivered', admin_name=? WHERE order_id=?", (admin_name, order_id))
    
    conn.commit()
    conn.close()
    
    user_msg = (
        f"🎉 **আপনার VPN অর্ডার ডেলিভারি সম্পন্ন হয়েছে!** 🎉\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"{emoji} **ব্র্যান্ড:** {vpn_name}\n"
        f"📦 **প্যাকেজ:** {duration}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"🔐 **আপনার লগইন ডিটেইলস:**\n"
        f"```text\n{vpn_info}\n```\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"💡 *(কপি করতে বক্সের উপর ক্লিক করুন)*"
    )
    
    try:
        await bot.send_message(user_id, user_msg, parse_mode="Markdown")
        await m.answer(f"✅ Successfully delivered to user!")
        try:
            await bot.edit_message_reply_markup(chat_id=m.chat.id, message_id=msg_id, reply_markup=None)
            await bot.edit_message_text(f"✅ **Delivered by {admin_name}**", chat_id=m.chat.id, message_id=msg_id)
        except: pass
    except Exception as e:
        await m.answer(f"⚠️ User delivery failed (Maybe blocked bot). DB updated. Error: {e}")
        
    await state.clear()

# --- NEW: SUPPORT SYSTEM & TERMS & POLICY ---

@dp.message(Command("terms"))
@dp.message(Command("rules"))
@dp.callback_query(F.data == "terms_policy")
async def show_terms_policy(event: types.Message | types.CallbackQuery):
    if isinstance(event, types.CallbackQuery):
        await event.answer()
        sender = event.message.edit_text
    else:
        sender = event.answer
        
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="🔄 রিপ্লেস ক্লেইম করুন", callback_data="sup_replace"))
    kb.row(types.InlineKeyboardButton(text="🔙 Back", callback_data="support_menu"))
    
    terms_text = (
        f"📜 **TERMS & POLICY • {BOT_VERSION} PREMIUM** 📜\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "আমাদের স্টোর থেকে কেনাকাটার পূর্বে ও পরে নিচের নীতিমালাগুলো বাধ্যতামূলকভাবে মেনে চলতে হবে:\n\n"
        "⏱️ **ডায়নামিক রিপ্লেস সময়সীমা (Replacement Tiers):**\n"
        "▫️ **১ – ৪ পিস আইডি:** ক্রয়ের পর **২ ঘণ্টা** গ্যারান্টি\n"
        "▫️ **৫ – ৩০ পিস আইডি:** ক্রয়ের পর **৬ ঘণ্টা** গ্যারান্টি\n"
        "▫️ **৩১ – ১০০ পিস আইডি:** ক্রয়ের পর **১২ ঘণ্টা** গ্যারান্টি\n"
        "▫️ **১০১+ পিস আইডি:** ক্রয়ের পর **২৪ ঘণ্টা** গ্যারান্টি\n\n"
        "🚫 *অনুমোদিত সময় পার হয়ে গেলে স্বয়ংক্রিয়ভাবে রিপ্লেস মেয়াদ শেষ হবে এবং কোনো রিকোয়েস্ট গ্রহণ করা হবে না।*\n\n"
        "📢 **PC Clone ID মার্কেট বাস্তবতা ও গুরুত্বপূর্ণ নির্দেশনা (বাধ্যতামূলক পাঠ্য):**\n"
        "▫️ আপনারা জানেন বর্তমানে মার্কেটে ফ্রেশ PC Clone ID পাওয়া অনেক কঠিন এবং দিন দিন তীব্র শর্টেজ দেখা দিচ্ছে। দীর্ঘ সময় ধরে ক্লোন আইডি সরবরাহ হওয়ায় মার্কেট সংকট সৃষ্টি হয়েছে।\n"
        "▫️ কিছু সেলার বর্তমানে ২-৩ মাস বা ১-২ সপ্তাহ আগের ব্যবহৃত (Used) আইডি পুনরায় বিভিন্ন বায়ারের কাছে বিক্রি করে থাকে। আমরা সর্বোচ্চ যাচাই করে বিশ্বস্ত সোর্স থেকে আইডি স্টকে নিই এবং সমস্যার প্রমাণ পেলে সেলারদের বরখাস্ত করি। কিন্তু চরম সংকটের সময় মার্কেটে এই বাস্তবতা উপেক্ষা করা যায় না।\n"
        "▫️ **ক্রেতাদের জন্য জরুরি পরামর্শ:** আইডি কেনার পর দ্রুত কাজ করে আপনার Ad Account, মাদার একাউন্ট বা অন্য সুরক্ষিত একাউন্টে ট্রান্সফার/মুভ করে রেখে দিন।\n"
        "▫️ **কঠোর নিয়ম:** আইডি ক্রয়ের জন্য নির্ধারিত গ্যারান্টি উইন্ডোর (২h / ৬h / ১২h / ২৪h) পর সেলারের রি-সেলিং বা লগইন সংক্রান্ত কোনো অভিযোগ গ্রহণযোগ্য হবে না। অতএব কেনার পরপরই আইডি চেক করে কাজ সম্পন্ন করুন।\n\n"
        "📌 **রিপ্লেসের নিয়মাবলী ও শর্তাবলী:**\n"
        "১. শুধুমাত্র `UID PASSWORD COOKIES` টেক্সট আকারে সাবমিট করতে হবে।\n"
        "২. কোনো ফটো, স্ক্রিনশট বা ফাইল পাঠালে রিকোয়েস্ট সাথে সাথে বাতিল হবে।\n"
        "৩. আইডি কেনার পর পাসওয়ার্ড পরিবর্তন করা হলে অথবা ২-ফ্যাক্টর অন করা হলে রিপ্লেস হবে না।\n"
        "৪. নিজস্ব আইপি বা ব্রাউজারের ভুলের কারণে আইডি নষ্ট হলে কিংবা আইডি লাইভ থাকলে রিপ্লেস হবে না।\n"
        "৫. একাধিক আইডি থাকলে প্রতি লাইনে একটি করে আইডি সাবমিট করতে হবে।\n\n"
        "💳 **ডিপোজিট ও রিফান্ড নীতি:**\n"
        "▫️ ব্যালেন্স রিফান্ডেবল নয়; যেকোনো সময় স্টোরের কেনাকাটায় ব্যবহার্য।\n"
        "▫️ ZiniPay বা Binance পেমেন্ট সম্পন্ন হওয়ার কয়েক সেকেন্ডে ব্যালেন্স অটো যোগ হয়।\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"🛡️ *BasicTrick Automated Store • {BOT_VERSION} Active*"
    )
    try:
        await sender(terms_text, reply_markup=kb.as_markup(), parse_mode="Markdown")
    except Exception:
        if isinstance(event, types.CallbackQuery):
            await event.message.answer(terms_text, reply_markup=kb.as_markup(), parse_mode="Markdown")

@dp.callback_query(F.data == "support_menu")
async def support_menu(c: types.CallbackQuery):
    await c.answer()
    
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="🔄 Replace PC Clone ID", callback_data="sup_replace"))
    kb.row(types.InlineKeyboardButton(text="📝 Complain", callback_data="sup_complain"))
    kb.row(types.InlineKeyboardButton(text="📜 শর্তাবলী ও পলিসি (Terms)", callback_data="terms_policy"))
    kb.row(types.InlineKeyboardButton(text="🔙 Back", callback_data="back_home"))
    
    msg = (
        f"📞 **সাপোর্ট ও হেল্প সেন্টার • {BOT_VERSION}**\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "আপনার যেকোনো সমস্যা বা সার্ভিসের জন্য নিচের অপশন সিলেক্ট করুন:\n\n"
        "💡 *নষ্ট আইডি রিপ্লেসের জন্য 'Replace PC Clone ID' এ ক্লিক করুন।*"
    )
    await c.message.edit_text(msg, reply_markup=kb.as_markup(), parse_mode="Markdown")

@dp.callback_query(F.data == "sup_replace")
async def support_replace_start(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    await state.clear()
    
    if not is_service_enabled("replace_service_enabled") and not is_admin(c.from_user.id):
        return await c.message.answer(
            "🔄 **রিপ্লেস রিকোয়েস্ট সার্ভিস সাময়িকভাবে বন্ধ আছে।**\n"
            "⏳ অনুগ্রহ করে কিছুক্ষণ পর আবার চেষ্টা করুন অথবা সরাসরি এডমিনের সাথে যোগাযোগ করুন।",
            parse_mode="Markdown"
        )
    
    conn = _dbc()
    user_sales = conn.execute("""
        SELECT id, category, qty, total, date, time 
        FROM sales 
        WHERE user_id=? AND category NOT LIKE 'VPN%'
        ORDER BY id DESC LIMIT 6
    """, (c.from_user.id,)).fetchall()
    conn.close()
    
    if not user_sales:
        kb = InlineKeyboardBuilder()
        kb.row(types.InlineKeyboardButton(text="📜 Terms & Policy", callback_data="terms_policy"))
        kb.row(types.InlineKeyboardButton(text="🔙 Back", callback_data="support_menu"))
        
        no_order_msg = (
            "🚫 **কোনো সক্রিয় অর্ডার পাওয়া যায়নি!**\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "⚠️ আপনার অ্যাকাউন্টে পূর্বে ক্রয়কৃত কোনো আইডি পাওয়া যায়নি।\n\n"
            "📌 **পলিসি নির্দেশিকা:**\n"
            "▫️ শুধুমাত্র আমাদের স্টোর থেকে সরাসরি ক্রয়কৃত আইডির ক্ষেত্রেই অটো রিপ্লেস গ্যারান্টি প্রযোজ্য।\n"
            "▫️ আপনি যদি অন্য টেলিগ্রাম অ্যাকাউন্ট থেকে কিনে থাকেন, অনুগ্রহ করে সেই অ্যাকাউন্ট থেকে যোগাযোগ করুন।"
        )
        return await c.message.edit_text(no_order_msg, reply_markup=kb.as_markup(), parse_mode="Markdown")
        
    kb = InlineKeyboardBuilder()
    now_ts = int(__import__("time").time())
    
    for sale in user_sales:
        s_id, s_cat, s_qty, s_tot, s_date, s_time = sale
        lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID","ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}.get(s_cat, s_cat.upper())
        
        s_epoch = get_sale_epoch(s_id, s_date, s_time)
        allowed_h = get_replace_window_hours(s_qty)
        is_expired = (now_ts - s_epoch) > (allowed_h * 3600)
        
        status_tag = "🔴 Expired" if is_expired else "🟢 Active"
        btn_text = f"📦 #{s_id} • {lbl} ({s_qty} pcs) [{status_tag}]"
        kb.row(types.InlineKeyboardButton(text=btn_text, callback_data=f"rep_ord_{s_id}"))
        
    kb.row(types.InlineKeyboardButton(text="📜 শর্তাবলী ও নিয়ম (Terms)", callback_data="terms_policy"))
    kb.row(types.InlineKeyboardButton(text="🔙 Back", callback_data="support_menu"))
    
    pick_msg = (
        f"🔄 **রিপ্লেস রিকোয়েস্ট সেন্টার • {BOT_VERSION}**\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "কোন অর্ডারের নষ্ট আইডি আপনি রিপ্লেস করতে চান তা নিচে থেকে সিলেক্ট করুন:\n\n"
        "⏱️ **আমাদের অটোমেটিক রিপ্লেস গ্যারান্টি:**\n"
        "▫️ ১ – ৪ পিস: **২ ঘণ্টা** গ্যারান্টি\n"
        "▫️ ৫ – ৩০ পিস: **৬ ঘণ্টা** গ্যারান্টি\n"
        "▫️ ৩১ – ১০০ পিস: **১২ ঘণ্টা** গ্যারান্টি\n"
        "▫️ ১০১+ পিস: **২৪ ঘণ্টা** গ্যারান্টি\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "💡 *সিস্টেম স্বয়ংক্রিয়ভাবে অর্ডারের সময় এবং পরিমাণের ওপর ভিত্তি করে রিপ্লেসের মেয়াদ যাচাই করবে।*"
    )
    await c.message.edit_text(pick_msg, reply_markup=kb.as_markup(), parse_mode="Markdown")

@dp.callback_query(F.data.startswith("rep_ord_"))
async def select_replace_order(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    s_id = int(c.data.split("_")[2])
    
    conn = _dbc()
    sale = conn.execute("SELECT id, user_id, username, category, qty, total, date, time FROM sales WHERE id=? AND user_id=?", (s_id, c.from_user.id)).fetchone()
    conn.close()
    
    if not sale:
        return await c.message.answer("❌ অর্ডারটি পাওয়া যায়নি।")
        
    sale_id, u_id, uname, cat_name, qty, total, d_str, t_str = sale
    lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID","ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}.get(cat_name, cat_name.upper())
    
    sale_epoch = get_sale_epoch(sale_id, d_str, t_str)
    now_ts = int(__import__("time").time())
    elapsed_sec = max(0, now_ts - sale_epoch)
    allowed_h = get_replace_window_hours(qty)
    allowed_sec = allowed_h * 3600
    
    elapsed_str = format_duration(elapsed_sec)
    
    if elapsed_sec > allowed_sec:
        # EXPIRED WARNING
        kb = InlineKeyboardBuilder()
        kb.row(types.InlineKeyboardButton(text="📜 Terms & Policy (শর্তাবলী)", callback_data="terms_policy"))
        kb.row(types.InlineKeyboardButton(text="🔙 অন্য অর্ডার বাছুন", callback_data="sup_replace"))
        
        expired_msg = (
            "🚫 **রিপ্লেস সময়সীমা অতিক্রম করেছে (Time Expired)** 🚫\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"📦 **অর্ডার নং:** `#{sale_id}`\n"
            f"🏷️ **আইটেম:** {lbl} ({qty} pcs)\n"
            f"🕒 **কেনার সময়:** {d_str} | {t_str}\n"
            f"⏳ **অনুমোদিত রিপ্লেস উইন্ডো:** {allowed_h} ঘণ্টা\n"
            f"⌛ **অতিবাহিত সময়:** {elapsed_str}\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            f"⚠️ **দুঃখিত!** আপনার অর্ডারের জন্য নির্ধারিত **{allowed_h} ঘণ্টার** রিপ্লেস সময়সীমা শেষ হয়ে গেছে।\n\n"
            "📜 আমাদের অটোমেটিক সিকিউরিটি ও টার্মস পলিসি অনুযায়ী নির্ধারিত সময় পার হওয়ার পর সিস্টেম থেকে কোনো রিপ্লেস গ্রহণ করা সম্ভব নয়।"
        )
        return await c.message.edit_text(expired_msg, reply_markup=kb.as_markup(), parse_mode="Markdown")
        
    # STILL VALID!
    rem_sec = allowed_sec - elapsed_sec
    rem_str = format_duration(rem_sec)
    
    await state.update_data(
        replace_sale_id=sale_id,
        replace_qty=qty,
        replace_cat=cat_name,
        replace_allowed_hours=allowed_h,
        replace_sale_time=f"{d_str} {t_str}",
        replace_sale_epoch=sale_epoch
    )
    
    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="✅ Agree & Submit", callback_data=f"rep_agree_{sale_id}"))
    kb.row(types.InlineKeyboardButton(text="📜 Terms & Policy", callback_data="terms_policy"))
    kb.row(types.InlineKeyboardButton(text="🔙 অন্য অর্ডার বাছুন", callback_data="sup_replace"))
    
    valid_msg = (
        "⚡ **অর্ডার ভেরিফাইড (রিপ্লেসের জন্য যোগ্য)** ⚡\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 **অর্ডার নং:** `#{sale_id}`\n"
        f"🏷️ **আইটেম:** {lbl} ({qty} pcs)\n"
        f"🕒 **কেনার সময়:** {d_str} | {t_str}\n"
        f"⏳ **অনুমোদিত গ্যারান্টি:** {allowed_h} ঘণ্টা\n"
        f"⏱️ **বাকি সময় আছে:** {rem_str}\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "📌 **নিয়মাবলী ও শর্তসমূহ:**\n"
        "১. শুধুমাত্র নষ্ট আইডি দিন (UID PASS COOKIES)।\n"
        "২. কোনো ফাইল, ফটো বা স্ক্রিনশট পাঠালে সরাসরি রিজেক্ট হবে।\n"
        "৩. পাসওয়ার্ড চেঞ্জ করা বা নিজস্ব ভুলের কারণে নষ্ট আইডি রিপ্লেস হবে না।\n\n"
        "✅ আপনি কি এই শর্তে নষ্ট আইডি সাবমিট করতে প্রস্তুত?"
    )
    await c.message.edit_text(valid_msg, reply_markup=kb.as_markup(), parse_mode="Markdown")

@dp.callback_query(F.data.startswith("rep_agree_"))
async def rep_agree_action(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    sale_id = c.data.split("_")[2]
    
    await c.message.edit_text(
        f"✍️ **অর্ডার `#{sale_id}` এর নষ্ট আইডিগুলো টেক্সট আকারে দিন:**\n\n"
        "📌 **Format (প্রতি লাইনে):**\n"
        "`UID PASSWORD COOKIES`\n\n"
        "📝 **Example:**\n"
        "`100011... myPass123 datr=xxx; c_user=100011...; xs=...`\n\n"
        "⚠️ শুধুমাত্র Text accept হবে। File / Screenshot পাঠালে auto-reject।\n"
        "একাধিক আইডি হলে প্রতি লাইনে একটি করে দিন।"
    )
    await state.set_state(ShopStates.waiting_for_replace_data)

@dp.message(ShopStates.waiting_for_replace_data)
async def process_replace_request(m: types.Message, state: FSMContext):
    # [REPLACE_TEXTONLY_PATCH_V1]
    if m.content_type != "text" or not m.text:
        return await m.answer(
            "❌ **File / Photo / Screenshot / Document accept হবে না।**\n\n"
            "📝 শুধু **text** paste করুন এই format-এ:\n"
            "`UID PASSWORD COOKIES`\n\n"
            "আবার চেষ্টা করুন অথবা /cancel দিন।"
        )

    if m.text.startswith("/"): return

    user_data_text = m.text.strip()

    # ---- Layer 2: per-line format validation ----
    raw_lines = [ln.strip() for ln in user_data_text.split("\n") if ln.strip()]
    if not raw_lines:
        return await m.answer("❌ খালি text। UID PASS COOKIES format-এ দিন।")

    _errors = []
    for _i, _ln in enumerate(raw_lines, 1):
        _parts = _ln.split(None, 2)
        if len(_parts) < 3:
            _errors.append(f"  • Line {_i}: শুধু {len(_parts)}টা field পাওয়া গেছে (দরকার 3 — UID PASS COOKIES)")
            continue
        _uid, _pw, _ck = _parts
        if len(_uid) < 3:
            _errors.append(f"  • Line {_i}: UID খুব ছোট ({_uid!r})")
        if len(_pw) < 3:
            _errors.append(f"  • Line {_i}: PASSWORD খুব ছোট")
        if "=" not in _ck:
            _errors.append(f"  • Line {_i}: COOKIES format ভুল (কোনো `=` নেই, যেমন `datr=...; c_user=...`)")

    if _errors:
        _err_txt = "\n".join(_errors[:8])
        _more = f"\n  ... আরো {len(_errors)-8} টি ত্রুটি" if len(_errors) > 8 else ""
        return await m.answer(
            f"⚠️ **আইডি ফরম্যাট সঠিক নয়! ({len(_errors)}টি লাইনে ত্রুটি)**\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"{_err_txt}{_more}\n\n"
            f"📌 **সঠিক ফরম্যাট দেওয়ার নিয়ম:**\n"
            f"প্রতিটি লাইনে ৩টি অংশ স্পেস (space) দিয়ে আলাদা করে থাকতে হবে:\n"
            f"`UID PASSWORD COOKIES`\n\n"
            f"💡 **নমুনা উদাহরণ (Example):**\n"
            f"```text\n"
            f"61593898545651 Saida@25 datr=3O2NaVk3u2E1lRsGtMNRU48O; c_user=61593898545651; xs=42%3A...\n"
            f"100029977223779 Juwel@3 datr=FAgZakFSQz1hOc15ekkIJlh9; c_user=100029977223779; xs=28%3A...\n"
            f"```\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"✍️ অনুগ্রহ করে ফরম্যাট ঠিক করে পুনরায় টেক্সট আকারে মেসেজ পাঠান অথবা বাতিল করতে /cancel লিখুন।",
            parse_mode="Markdown"
        )

    # Re-verify time expiration in state
    st_data = await state.get_data()
    sale_id = st_data.get("replace_sale_id")
    allowed_h = st_data.get("replace_allowed_hours", 6)
    sale_epoch = st_data.get("replace_sale_epoch")
    qty = st_data.get("replace_qty", len(raw_lines))

    now_ts = int(__import__("time").time())
    if sale_epoch and (now_ts - sale_epoch) > (allowed_h * 3600):
        await state.clear()
        kb = InlineKeyboardBuilder()
        kb.row(types.InlineKeyboardButton(text="📜 Terms & Policy", callback_data="terms_policy"))
        kb.row(types.InlineKeyboardButton(text="🔙 Back", callback_data="support_menu"))
        return await m.answer(
            f"🚫 **রিপ্লেস সময়সীমা অতিক্রম করেছে!**\n\nআইডি সাবমিট করার আগেই আপনার অর্ডারের {allowed_h} ঘণ্টার গ্যারান্টি মেয়াদ শেষ হয়ে গেছে।",
            reply_markup=kb.as_markup()
        )

    ticket_id = str(uuid.uuid4())[:8]
    order_ref = f"Order #{sale_id}" if sale_id else "Direct"
    lines = raw_lines
    acc_count_warning = f"⚠️ __ইউজার {len(lines)} টি একাউন্ট দিয়েছে!__" if len(lines) > 1 else ""
    
    username_display = f"@{m.from_user.username}" if m.from_user.username else "No Username"
    utc_now_ts = datetime.now(timezone.utc).timestamp()
    
    conn = _dbc()
    try:
        conn.execute(
            "INSERT INTO support_tickets (ticket_id, user_id, type, status, data, timestamp) VALUES (?, ?, 'replace', 'pending', ?, ?)",
            (ticket_id, m.from_user.id, user_data_text, utc_now_ts)
        )
        _rep_ts = int(utc_now_ts * 1000)
        _rep_uname = f"@{m.from_user.username}" if m.from_user.username else (m.from_user.first_name or f"User_{m.from_user.id}")
        conn.execute(
            "INSERT INTO replace_requests (user_id, username, category, old_data, reason, status, created_at) VALUES (?, ?, ?, ?, ?, 'pending', ?)",
            (m.from_user.id, _rep_uname, "ID", user_data_text, f"Ticket #{ticket_id} ({order_ref})", _rep_ts)
        )
        conn.commit()
    except Exception as e:
        print(f"[replace] insert error: {e}", flush=True)
        await m.answer("❌ Error processing request.")
        conn.close()
        return
        
    admins = conn.execute("SELECT user_id FROM admins").fetchall()
    conn.close()
    
    short_text = user_data_text[:60] + "..." if len(user_data_text) > 60 else user_data_text
    
    admin_msg = (
        f"🚨 **NEW REPLACE REQUEST • {BOT_VERSION}** 🚨\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 **{order_ref}** ({qty} pcs | Tier: {allowed_h}h)\n"
        f"👤 **Name:** {m.from_user.first_name}\n"
        f"🔗 **User:** {username_display}\n"
        f"🆔 **ID:** `{m.from_user.id}`\n\n"
        f"📋 **Provided Details:**\n`{short_text}`\n\n"
        f"{acc_count_warning}"
    )
    
    kb = InlineKeyboardBuilder()
    if len(user_data_text) > 60:
        kb.row(types.InlineKeyboardButton(text="📄 See Full Details", callback_data=f"tick_view_{ticket_id}"))
    kb.row(types.InlineKeyboardButton(text="🔄 Replace", callback_data=f"tick_rep_{ticket_id}"))
    kb.row(types.InlineKeyboardButton(text=f"⏱ Time Over ({allowed_h}h)", callback_data=f"tick_timeover_{ticket_id}_{allowed_h}"))
    kb.row(types.InlineKeyboardButton(text="✉️ Reply", callback_data=f"tick_reply_{ticket_id}"))
    kb.row(types.InlineKeyboardButton(text="🔕 Cancel Reminder", callback_data=f"tick_cancelrem_{ticket_id}"))
    
    for admin in admins:
        try: await bot.send_message(admin[0], admin_msg, reply_markup=kb.as_markup())
        except: pass
        
    await m.answer(
        f"✅ **আপনার রিকোয়েস্টটি অ্যাডমিনের কাছে সফলভাবে পাঠানো হয়েছে!**\n\n"
        f"📦 **অর্ডার:** {order_ref}\n"
        f"🎫 **টিকিট আইডি:** `#{ticket_id}`\n"
        f"অ্যাডমিন চেক করে দ্রুত রিপ্লেস দিয়ে দিবে।"
    )
    await state.clear()
    
    asyncio.create_task(ticket_reminder(ticket_id, admins, admin_msg, kb))

# COMPLAIN SYSTEM
@dp.callback_query(F.data == "sup_complain")
async def support_complain_input(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    await c.message.edit_text("✍️ **আপনার অভিযোগ বা সমস্যার কথা বিস্তারিত লিখুন:**")
    await state.set_state(ShopStates.waiting_for_complain_text)

@dp.message(ShopStates.waiting_for_complain_text, F.content_type.in_({"text", "photo", "document"}))
async def process_complain_request(m: types.Message, state: FSMContext):
    if m.text and m.text.startswith("/"): return

    complain_text = (m.text or m.caption or "").strip()
    photo_id = m.photo[-1].file_id if m.photo else None
    doc_id = m.document.file_id if (m.document and (m.document.mime_type or "").startswith("image/")) else None

    if not complain_text and not photo_id and not doc_id:
        await m.answer("❌ অনুগ্রহ করে সমস্যার বিবরণ লিখুন (ছবি দিলে caption-এ লিখুন)।")
        return

    ticket_id = str(uuid.uuid4())[:8]
    username_display = f"@{m.from_user.username}" if m.from_user.username else "No Username"

    conn = _dbc()
    try:
        conn.execute("INSERT INTO support_tickets (ticket_id, user_id, type, status) VALUES (?, ?, 'complain', 'pending')", (ticket_id, m.from_user.id))
        conn.commit()
        admins = conn.execute("SELECT user_id FROM admins").fetchall()
    except Exception:
        conn.close()
        await m.answer("❌ Error processing request.")
        return
    conn.close()

    has_img = "🖼 Screenshot: যুক্ত আছে" if (photo_id or doc_id) else "🖼 Screenshot: নেই"
    admin_msg = (
        f"📝 **NEW COMPLAIN / MSG** 📝\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 **Name:** {m.from_user.first_name}\n"
        f"🔗 **User:** {username_display}\n"
        f"🆔 **ID:** `{m.from_user.id}`\n"
        f"🎫 **Ticket:** `{ticket_id}`\n"
        f"{has_img}\n\n"
        f"💬 **Message:**\n{complain_text or '(শুধু ছবি পাঠানো হয়েছে)'}"
    )

    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="✉️ Reply", callback_data=f"tick_reply_{ticket_id}"))

    for admin in admins:
        try:
            if photo_id:
                await bot.send_photo(admin[0], photo_id, caption=admin_msg, reply_markup=kb.as_markup())
            elif doc_id:
                await bot.send_document(admin[0], doc_id, caption=admin_msg, reply_markup=kb.as_markup())
            else:
                await bot.send_message(admin[0], admin_msg, reply_markup=kb.as_markup())
        except Exception:
            pass

    await m.answer("✅ **আপনার মেসেজটি অ্যাডমিনের কাছে পাঠানো হয়েছে!**\nঅ্যাডমিন দ্রুত আপনাকে রিপ্লাই করবে।")
    await state.clear()

@dp.message(ShopStates.waiting_for_complain_text)
async def process_complain_unsupported(m: types.Message):
    await m.answer("⚠️ শুধু **টেক্সট** অথবা **স্ক্রিনশট (ছবি)** পাঠানো যাবে।")

# CUSTOM REMINDER TASK
async def ticket_reminder(ticket_id, admins, original_msg, kb):
    intervals = [1500, 3600, 7200, 10800] 
    
    for wait_time in intervals:
        await asyncio.sleep(wait_time)
        conn = _dbc()
        status = conn.execute("SELECT status FROM support_tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
        conn.close()
        
        if not status or status[0] != 'pending':
            break 
            
        rem_msg = f"⏰ **REMINDER! PENDING REPLACE REQUEST:**\n\n{original_msg}"
        for admin in admins:
            try: await bot.send_message(admin[0], rem_msg, reply_markup=kb.as_markup())
            except: pass

# ADMIN TICKET ACTIONS

@dp.callback_query(F.data.startswith("tick_view_"))
async def tick_view_action(c: types.CallbackQuery):
    await c.answer()
    ticket_id = c.data.split("_")[2]
    conn = _dbc()
    ticket = conn.execute("SELECT user_id, data FROM support_tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
    conn.close()
    
    if ticket:
        full_text = ticket[1] or ""
        kb = InlineKeyboardBuilder()
        kb.row(types.InlineKeyboardButton(text="🔄 Replace", callback_data=f"tick_rep_{ticket_id}"))
        kb.row(types.InlineKeyboardButton(text="⏱ Time Over", callback_data=f"tick_timeover_{ticket_id}_6"))
        kb.row(types.InlineKeyboardButton(text="✉️ Reply", callback_data=f"tick_reply_{ticket_id}"))
        kb.row(types.InlineKeyboardButton(text="🔕 Cancel Reminder", callback_data=f"tick_cancelrem_{ticket_id}"))

        if len(full_text) > 2000:
            doc = BufferedInputFile(full_text.encode('utf-8'), filename=f"replace_details_{ticket_id}.txt")
            try:
                await c.message.answer_document(
                    doc,
                    caption=f"📄 **Full Replace Details (Ticket #{ticket_id})**\n🆔 **User:** `{ticket[0]}`\n📏 Size: {len(full_text)} characters",
                    reply_markup=kb.as_markup()
                )
            except Exception as e:
                print(f"[tick_view] doc error: {e}", flush=True)
                await c.message.answer(f"⚠️ Text too large to display directly: {len(full_text)} chars.", reply_markup=kb.as_markup())
        else:
            new_msg = (
                f"🚨 **REPLACE REQUEST (FULL)** 🚨\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"🆔 **ID:** `{ticket[0]}`\n\n"
                f"📋 **Full Details:**\n`{full_text}`\n"
            )
            try: await c.message.edit_text(new_msg, reply_markup=kb.as_markup())
            except Exception as e:
                print(f"[tick_view] edit_text error: {e}", flush=True)

@dp.callback_query(F.data.startswith("tick_cancelrem_"))
async def tick_cancel_reminder_action(c: types.CallbackQuery):
    await c.answer()
    ticket_id = c.data.split("_")[2]
    
    conn = _dbc()
    ticket = conn.execute("SELECT status FROM support_tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
    
    if not ticket or ticket[0] != 'pending':
        conn.close()
        try: await c.message.edit_reply_markup(reply_markup=None)
        except: pass
        return await c.message.reply("⚠️ Already processed or cancelled.")
        
    conn.execute("UPDATE support_tickets SET status='ignored' WHERE ticket_id=?", (ticket_id,))
    conn.commit()
    conn.close()
    
    try: await c.message.edit_reply_markup(reply_markup=None)
    except: pass
    try: await c.message.edit_text(f"{c.message.text}\n\n🔕 **Reminder Cancelled by {c.from_user.first_name}**")
    except: pass

@dp.callback_query(F.data.startswith("tick_timeover_"))
@dp.callback_query(F.data.startswith("tick_6hr_"))
async def tick_timeover_action(c: types.CallbackQuery):
    await c.answer()
    parts = c.data.split("_")
    ticket_id = parts[2]
    allowed_h = parts[3] if len(parts) > 3 else "6"
    
    conn = _dbc()
    ticket = conn.execute("SELECT user_id, status FROM support_tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
    
    if not ticket or (ticket[1] != 'pending' and ticket[1] != 'ignored'):
        conn.close()
        try: await c.message.edit_reply_markup(reply_markup=None)
        except: pass
        return await c.message.reply("⚠️ Already processed.")
        
    user_id = ticket[0]
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
    admin_name = f"bot-admin ({c.from_user.first_name})"
    conn.execute("UPDATE support_tickets SET status='processed', admin_response=? WHERE ticket_id=?", (f"Time Over ({allowed_h}h)", ticket_id))
    conn.execute("""
        UPDATE replace_requests 
        SET status='rejected', replacement_data=?, resolved_by=?, resolved_at=? 
        WHERE (reason LIKE ? OR (user_id=? AND status='pending'))
    """, (f"[Time Over: {allowed_h}h limit exceeded]", admin_name, now_ms, f"%{ticket_id}%", user_id))
    conn.commit()
    conn.close()
    
    try: await c.message.edit_reply_markup(reply_markup=None)
    except: pass
    try: await c.message.edit_text(f"{c.message.text}\n\n❌ **Rejected (Time Over - {allowed_h}h) by {c.from_user.first_name}**")
    except: pass
    
    user_msg = (
        f"😔 **দুঃখিত!**\n\n"
        f"আপনার অর্ডারের জন্য নির্ধারিত **{allowed_h} ঘণ্টার** রিপ্লেস সময়সীমা অতিক্রম করেছে।\n"
        f"আমাদের স্টোরের অটোমেটিক পলিসি ও রুলস অনুযায়ী সময়সীমা শেষ হওয়ার পর আইডি রিপ্লেস দেওয়া সম্ভব নয়।\n\n"
        f"পরবর্তীতে আইডি কেনার পর দ্রুত চেক করে কোনো সমস্যা থাকলে নির্দিষ্ট সময়ের মধ্যে জানানোর জন্য অনুরোধ করা হলো।"
    )
    try: await bot.send_message(user_id, user_msg)
    except: pass

@dp.callback_query(F.data.startswith("tick_reply_"))
async def tick_reply_action(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    ticket_id = c.data.split("_")[2]
    
    conn = _dbc()
    ticket = conn.execute("SELECT user_id, status FROM support_tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
    conn.close()
    
    if not ticket or (ticket[1] != 'pending' and ticket[1] != 'ignored'):
        try: await c.message.edit_reply_markup(reply_markup=None)
        except: pass
        return await c.message.reply("⚠️ Already processed.")
        
    await state.update_data(current_ticket_user=ticket[0], current_ticket_id=ticket_id, admin_msg_id=c.message.message_id)
    await c.message.reply("✍️ **ইউজারকে যে মেসেজ দিতে চান তা লিখুন:**")
    await state.set_state(ShopStates.waiting_for_admin_reply)

@dp.message(ShopStates.waiting_for_admin_reply)
async def send_admin_reply(m: types.Message, state: FSMContext):
    data = await state.get_data()
    user_id = data.get('current_ticket_user')
    ticket_id = data.get('current_ticket_id')
    msg_id = data.get('admin_msg_id')
    
    if not user_id: return await m.answer("❌ Error.")
    
    reply_text = m.text or ""
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
    admin_name = f"bot-admin ({m.from_user.first_name})"
    
    conn = _dbc()
    conn.execute("UPDATE support_tickets SET status='processed', admin_response=? WHERE ticket_id=?", (reply_text, ticket_id))
    # Sync with replace_requests so it does not stay in pending!
    conn.execute("""
        UPDATE replace_requests 
        SET status='replaced', replacement_data=?, resolved_by=?, resolved_at=? 
        WHERE (reason LIKE ? OR (user_id=? AND status='pending'))
    """, (f"[Replied: {reply_text}]", admin_name, now_ms, f"%{ticket_id}%", user_id))
    conn.commit()
    conn.close()
    
    user_msg = f"📩 **Admin Message:**\n━━━━━━━━━━━━━━━━━━━━\n{reply_text}"
    
    try:
        await bot.send_message(user_id, user_msg)
        await m.answer("✅ মেসেজ ইউজারের কাছে পাঠানো হয়েছে এবং রিকোয়েস্ট সমাধান হিসেবে মার্ক করা হয়েছে।")
        try: await bot.edit_message_text(f"✅ **Replied & Solved by {m.from_user.first_name}**", chat_id=m.chat.id, message_id=msg_id)
        except: pass
    except:
        await m.answer("❌ User delivery failed.")
        
    await state.clear()

@dp.callback_query(F.data.startswith("tick_rep_"))
async def tick_replace_action(c: types.CallbackQuery, state: FSMContext):
    await c.answer()
    ticket_id = c.data.split("_")[2]
    
    conn = _dbc()
    ticket = conn.execute("SELECT user_id, status FROM support_tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
    conn.close()
    
    if not ticket or (ticket[1] != 'pending' and ticket[1] != 'ignored'):
        try: await c.message.edit_reply_markup(reply_markup=None)
        except: pass
        return await c.message.reply("⚠️ Already processed.")
        
    await state.update_data(current_ticket_user=ticket[0], current_ticket_id=ticket_id, admin_msg_id=c.message.message_id)
    await c.message.reply("✍️ **নতুন রিপ্লেসমেন্ট আইডিগুলো দিন:**")
    await state.set_state(ShopStates.waiting_for_admin_replace)

@dp.message(ShopStates.waiting_for_admin_replace)
async def send_admin_replace(m: types.Message, state: FSMContext):
    data = await state.get_data()
    user_id = data.get('current_ticket_user')
    ticket_id = data.get('current_ticket_id')
    msg_id = data.get('admin_msg_id')
    
    if not user_id: return await m.answer("❌ Error.")
    
    replace_data = m.text or ""
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
    admin_name = f"bot-admin ({m.from_user.first_name})"
    
    conn = _dbc()
    ticket_data = conn.execute("SELECT data FROM support_tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
    original_data = ticket_data[0] if ticket_data else "N/A"
    
    conn.execute("UPDATE support_tickets SET status='processed', admin_response=? WHERE ticket_id=?", (replace_data, ticket_id))
    conn.execute("""
        UPDATE replace_requests 
        SET status='replaced', replacement_data=?, resolved_by=?, resolved_at=? 
        WHERE reason LIKE ? AND status='pending'
    """, (replace_data, admin_name, now_ms, f"%{ticket_id}%"))
    conn.commit()
    conn.close()
    
    user_delivered = False
    if len(replace_data) > 2500:
        doc = BufferedInputFile(replace_data.encode('utf-8'), filename=f"replacement_{ticket_id}.txt")
        try:
            await bot.send_document(
                user_id,
                doc,
                caption="✅ **আপনার রিপ্লেসমেন্ট আইডি দেওয়া হয়েছে!**\nফাইলে সম্পূর্ণ তথ্য দেখতে পাবেন।"
            )
            user_delivered = True
        except Exception as e:
            print(f"[send_admin_replace] send_document error: {e}", flush=True)
    else:
        user_msg = (
            f"✅ **আপনার রিপ্লেসমেন্ট আইডি দেওয়া হয়েছে:**\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"```text\n{replace_data}\n```\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"💡 *(কপি করতে বক্সের উপর ক্লিক করুন)*"
        )
        try:
            await bot.send_message(user_id, user_msg, parse_mode="Markdown")
            user_delivered = True
        except Exception as e:
            print(f"[send_admin_replace] send_message error: {e}", flush=True)
            
    if user_delivered:
        await m.answer("✅ রিপ্লেসমেন্ট আইডি ইউজারের কাছে পাঠানো হয়েছে।")
        
        orig_preview = original_data[:200] + "..." if len(original_data) > 200 else original_data
        rep_preview = replace_data[:200] + "..." if len(replace_data) > 200 else replace_data
        final_admin_msg = (
            f"🚨 **REPLACE REQUEST (RESOLVED)** 🚨\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 **User ID:** `{user_id}`\n\n"
            f"📋 **Original Details:**\n`{orig_preview}`\n\n"
            f"✅ **REPLACED BY {m.from_user.first_name}**\n"
            f"**New Data:** `{rep_preview}`"
        )
        try: await bot.edit_message_text(final_admin_msg, chat_id=m.chat.id, message_id=msg_id, reply_markup=None)
        except: pass
    else:
        await m.answer("❌ User delivery failed.")
        
    await state.clear()

@dp.callback_query(F.data == "profile")
async def profile(c: types.CallbackQuery):
    await c.answer()
    bal, _, _ = get_user_data(c.from_user.id, c.from_user.username, c.from_user.first_name)
    conn = _dbc()
    order_stat = conn.execute(
        "SELECT count(*), coalesce(sum(total),0) FROM sales WHERE user_id=?",
        (c.from_user.id,)
    ).fetchone()
    conn.close()

    total_orders = order_stat[0] if order_stat else 0
    total_spent = order_stat[1] if order_stat else 0

    kb = InlineKeyboardBuilder()
    kb.row(types.InlineKeyboardButton(text="📦 আমার অর্ডারসমূহ (Order History)", callback_data="my_orders"))
    kb.row(types.InlineKeyboardButton(text="🔙 ব্যাক", callback_data="back_home"))

    text = (
        f"👤 **ইউজার প্রোফাইল**\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 **আইডি:** `{c.from_user.id}`\n"
        f"👤 **নাম:** {c.from_user.first_name}\n"
        f"💰 **বর্তমান ব্যালেন্স:** `{bal}৳`\n"
        f"🛍️ **মোট অর্ডার:** `{total_orders}টি`\n"
        f"💸 **মোট খরচ:** `{total_spent}৳`\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"বিগত অর্ডারের ফাইল দেখতে নিচে 'আমার অর্ডারসমূহ' বাটনে চাপুন।"
    )
    await c.message.edit_text(text, reply_markup=kb.as_markup(), parse_mode="Markdown")


@dp.callback_query(F.data == "my_orders")
async def my_orders_handler(c: types.CallbackQuery):
    await c.answer()
    conn = _dbc()
    sales = conn.execute(
        "SELECT id, category, qty, total, date, time FROM sales WHERE user_id=? ORDER BY id DESC LIMIT 5",
        (c.from_user.id,)
    ).fetchall()
    conn.close()

    if not sales:
        kb = InlineKeyboardBuilder().row(types.InlineKeyboardButton(text="🔙 প্রোফাইলে ফিরে যান", callback_data="profile"))
        return await c.message.edit_text("📦 আপনার কোনো পূর্ববর্তী অর্ডার পাওয়া যায়নি।", reply_markup=kb.as_markup())

    kb = InlineKeyboardBuilder()
    lines = ["📦 **আপনার সাম্প্রতিক ৫টি অর্ডার:**\n━━━━━━━━━━━━━━━━━━━━"]
    for s in sales:
        sid, cat, qty, total, dt, tm = s
        lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID","ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}.get(cat, cat.upper())
        lines.append(f"• **Order #{sid}** | {lbl} × {qty} | `{total}৳`\n  📅 {dt} {tm or ''}")
        kb.row(
            types.InlineKeyboardButton(text=f"📊 #{sid} Excel", callback_data=f"dfmt:xlsx:{sid}"),
            types.InlineKeyboardButton(text=f"📝 #{sid} Text", callback_data=f"dfmt:txt:{sid}")
        )

    kb.row(types.InlineKeyboardButton(text="🔙 প্রোফাইলে ফিরে যান", callback_data="profile"))
    await c.message.edit_text("\n\n".join(lines), reply_markup=kb.as_markup(), parse_mode="Markdown")


# ===== /checkpay command (ZiniPay manual verify) =====
import os as _os_chk
import sqlite3 as _sql_chk
import urllib.request as _req_chk
import urllib.parse as _parse_chk
import json as _json_chk

# [BUYLIMIT_V1] ---- 10 pcs / 10 min limit (শুধু FB 1000xx) ----
BUYLIMIT_MAX    = 10
BUYLIMIT_WINDOW = 600  # seconds
# শুধু এই category গুলোতে limit; fb61 / tempid / অন্যসব unlimited
BUYLIMIT_CATS   = {"fb1000", "fb1000xx", "1000xx"}
# web panel (config table) থেকে on/off + value override — 5s cache
_BL_CFG_CACHE = {"t": 0, "v": {}}

def _bl_cfg():
    """config table থেকে buylimit_* key গুলো পড়ে (web panel controlled)"""
    import time as _t
    now = _t.time()
    if now - _BL_CFG_CACHE["t"] < 5:
        return _BL_CFG_CACHE["v"]
    vals = {}
    conn = None
    try:
        conn = _bl_conn()
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)")
        for k, v in cur.execute(
            "SELECT key, value FROM config WHERE key IN "
            "('buylimit_enabled','buylimit_max','buylimit_window_min','buylimit_cats')"
        ).fetchall():
            vals[str(k)] = str(v)
        conn.commit()
    except Exception:
        pass
    finally:
        if conn is not None:
            try: conn.close()
            except Exception: pass
    _BL_CFG_CACHE["t"] = now
    _BL_CFG_CACHE["v"] = vals
    return vals

def _bl_enabled():
    v = str(_bl_cfg().get("buylimit_enabled", "on")).strip().lower()
    return v not in ("0", "off", "false", "no", "disabled", "closed")

def _bl_max():
    try:
        n = int(str(_bl_cfg().get("buylimit_max", BUYLIMIT_MAX)).strip())
        return n if 1 <= n <= 100000 else BUYLIMIT_MAX
    except Exception:
        return BUYLIMIT_MAX

def _bl_window():
    try:
        n = int(str(_bl_cfg().get("buylimit_window_min", BUYLIMIT_WINDOW // 60)).strip())
        return n * 60 if 1 <= n <= 1440 else BUYLIMIT_WINDOW
    except Exception:
        return BUYLIMIT_WINDOW

def _bl_cats():
    raw = str(_bl_cfg().get("buylimit_cats", "")).strip()
    if not raw:
        return BUYLIMIT_CATS
    s = {c.strip().lower() for c in raw.replace("\n", ",").split(",") if c.strip()}
    return s or BUYLIMIT_CATS

def _bl_limited(cat):
    try:
        if not _bl_enabled():
            return False
        return str(cat or "").strip().lower() in _bl_cats()
    except Exception:
        return False


def _bl_db_path():
    _p = globals().get("DB_FILE") or globals().get("DB_PATH") or globals().get("DB")
    if isinstance(_p, str) and _p.strip():
        return _p
    import os as _os
    for _cand in ("/root/store.db", _os.path.join(_os.path.dirname(__file__), "store.db"), "store.db"):
        try:
            if _os.path.exists(_cand):
                return _cand
        except Exception:
            pass
    return "/root/store.db"

def _bl_conn():
    try:
        return _dbc()
    except Exception:
        import sqlite3 as _s
        c = _s.connect(_bl_db_path(), timeout=15)
        try:
            c.execute("PRAGMA busy_timeout=15000")
        except Exception:
            pass
        return c

def _bl_init(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS buy_limit (
        user_id INTEGER PRIMARY KEY,
        window_start INTEGER NOT NULL,
        count INTEGER NOT NULL DEFAULT 0)""")

def _bl_state(uid):
    """(used_in_window, seconds_left) — window expired হলে (0, 0)"""
    import time as _t
    now = int(_t.time())
    conn = _bl_conn()
    try:
        cur = conn.cursor()
        _bl_init(cur)
        row = cur.execute("SELECT window_start, count FROM buy_limit WHERE user_id=?", (uid,)).fetchone()
        conn.commit()
    finally:
        try: conn.close()
        except Exception: pass
    if not row:
        return 0, 0
    ws, cnt = int(row[0] or 0), int(row[1] or 0)
    if ws <= 0 or now - ws >= _bl_window():
        return 0, 0
    return cnt, _bl_window() - (now - ws)

def _bl_fmt_left(secs):
    secs = max(0, int(secs))
    return f"{secs // 60} মিনিট {secs % 60} সেকেন্ড"

def _bl_allow(uid, qty=1, cat=None):
    """(ok, used, left_secs, allowed) — per-pcs cumulative counting.
    • qty <= বাকি pcs  -> ok=True, allowed=qty
    • qty > বাকি pcs   -> ok=False, allowed=বাকি pcs (0 হলে পুরো block)
    • non-limited category (fb61/tempid) -> সবসময় ok, allowed=qty
    """
    try: qty = max(1, int(qty))
    except Exception: qty = 1
    if not _bl_limited(cat):
        return True, 0, 0, qty
    used, left = _bl_state(uid)
    remain = max(0, _bl_max() - used)
    if qty <= remain:
        return True, used, left, qty
    return False, used, left, remain

def _bl_commit(uid, qty=1, cat=None):
    """কেনার পরে count বাড়ায় (যত pcs কিনেছে ততই); ফেরত দেয় (used, left_secs)"""
    if not _bl_limited(cat):
        return 0, 0
    import time as _t
    now = int(_t.time())
    try: qty = max(1, int(qty))
    except Exception: qty = 1
    conn = _bl_conn()
    try:
        cur = conn.cursor()
        _bl_init(cur)
        row = cur.execute("SELECT window_start, count FROM buy_limit WHERE user_id=?", (uid,)).fetchone()
        if row and int(row[0] or 0) > 0 and now - int(row[0]) < _bl_window():
            ws, cnt = int(row[0]), int(row[1] or 0) + qty
            cur.execute("UPDATE buy_limit SET count=? WHERE user_id=?", (cnt, uid))
        else:
            ws, cnt = now, qty
            cur.execute(
                "INSERT INTO buy_limit (user_id, window_start, count) VALUES (?,?,?) "
                "ON CONFLICT(user_id) DO UPDATE SET window_start=excluded.window_start, count=excluded.count",
                (uid, ws, cnt))
        conn.commit()
    finally:
        try: conn.close()
        except Exception: pass
    return cnt, max(0, _bl_window() - (now - ws))

def _bl_block_text(used, left, want=None, allowed=0):
    used = max(0, int(used or 0))
    remain = max(0, _bl_max() - used)
    if remain > 0:
        head = (
            "⚠️ **লিমিটের বেশি চাওয়া হয়েছে**\n\n"
            f"🧾 এই উইন্ডোতে ব্যবহার: **{used}/{_bl_max()} pcs**\n"
            f"✅ এখন সর্বোচ্চ নিতে পারবেন: **{remain} pcs**\n"
            f"⏳ পুরো লিমিট রিসেট হবে: **{_bl_fmt_left(left)}** পরে\n\n"
        )
    else:
        head = (
            "⛔ **কেনার লিমিট শেষ**\n\n"
            f"🧾 আপনি এই উইন্ডোতে **{used}/{_bl_max()} pcs** নিয়ে ফেলেছেন।\n"
            f"⏳ আবার নিতে পারবেন: **{_bl_fmt_left(left)}** পরে\n\n"
        )
    return head + (
        f"ℹ️ নিয়ম: **FB 1000xx** এর জন্য প্রতি **{_bl_window()//60} মিনিটে সর্বোচ্চ {_bl_max()} pcs** "
        "(কম কম করে নিলেও যোগ হয়ে হিসাব হবে)। FB 61 ও Temp ID unlimited।"
    )

def _bl_ok_text(used, left):
    used = max(0, int(used or 0))
    remain = max(0, _bl_max() - used)
    if remain <= 0:
        return (
            f"⏱ **লিমিট পূর্ণ:** এই উইন্ডোতে **{used}/{_bl_max()} pcs** শেষ\n"
            f"🔄 নতুন {_bl_max()} pcs লিমিট চালু হবে **{_bl_fmt_left(left)}** পরে"
        )
    return (
        f"⏱ **লিমিট আপডেট:** এই উইন্ডোতে **{used}/{_bl_max()} pcs** ব্যবহার হয়েছে "
        f"(বাকি **{remain} pcs**)\n"
        f"🔄 রিসেট হবে **{_bl_fmt_left(left)}** পরে"
    )
# [BUYLIMIT_V1] ---- end helper ----


_VPS_URL_CHK = _os_chk.environ.get("VPS_ADMIN_URL", "http://127.0.0.1:3000")
_DL_SECRET_CHK = _os_chk.environ.get("DOWNLOAD_SECRET", "")
_DB_PATH_CHK = _os_chk.environ.get("STORE_DB", "/root/store.db")

try:
    from aiogram.filters import Command as _CmdChk

    @dp.message(_CmdChk("checkpay"))
    async def _checkpay_handler(m):
        uid = m.from_user.id

        try:
            con = _sql_chk.connect(_DB_PATH_CHK)
            cur = con.cursor()
            cur.execute(
                "SELECT req_id, amount FROM payment_logs "
                "WHERE user_id=? AND method='zinipay' AND status='pending' "
                "ORDER BY timestamp DESC LIMIT 10",
                (uid,)
            )
            rows = cur.fetchall()
            con.close()
        except Exception as e:
            await m.answer(f"❌ DB error: {e}")
            return

        if not rows:
            await m.answer("কোন pending ZiniPay payment পাওয়া যায়নি।")
            return

        msg = await m.answer("⏳ Pending payment check করছি...")

        approved = 0
        pending = 0
        failed = 0
        lines = []

        for req_id, amount in rows:
            try:
                url = f"{_VPS_URL_CHK}/zinipay/check/{_parse_chk.quote(str(req_id))}?secret={_parse_chk.quote(_DL_SECRET_CHK)}"
                with _req_chk.urlopen(url, timeout=15) as r:
                    data = _json_chk.loads(r.read().decode())

                status = str(data.get("status", "")).lower()
                txid = data.get("transaction_id") or "-"

                if status == "approved":
                    approved += 1
                    icon = "✅"
                elif status == "pending":
                    pending += 1
                    icon = "⏳"
                else:
                    failed += 1
                    icon = "❌"

                lines.append(f"{icon} {req_id} | ৳{amount} | {status} | {txid}")
            except Exception as e:
                failed += 1
                lines.append(f"❌ {req_id} | ৳{amount} | error: {e}")

        text = (
            "💳 ZiniPay Payment Status\n\n"
            f"✅ Approved: {approved}\n"
            f"⏳ Pending (auto-approve within 60s): {pending + failed}\n\n"
            "ℹ️ Pending থাকলে চিন্তা করবেন না — payment পেলে auto balance যোগ হবে।\n\n"
            + "\n".join(lines[:10])
        )

        await msg.edit_text(text)

except Exception as _e_chk:
    print("checkpay handler load failed:", _e_chk)


async def main():
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())

