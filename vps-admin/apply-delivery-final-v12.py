#!/usr/bin/env python3
"""V12: fix TXT-vs-XLSX mixup + keep buttons so the same file can be re-downloaded."""
import os
import shutil
import sys
import time

STORE = "/root/store.py"
if not os.path.isfile(STORE):
    print(f"❌ Missing {STORE}")
    sys.exit(1)

src = open(STORE, encoding="utf-8").read()
markers = (
    "# [DELIVERY_CALLBACK_FIX_V4]",
    "# [DELIVERY_CALLBACK_FIX_V5]",
    "# [DELIVERY_UPLOAD_FIX_V6]",
    "# [DELIVERY_FINAL_V11]",
    "# [DELIVERY_FINAL_V12]",
)
starts = [src.find(m) for m in markers if src.find(m) >= 0]
if not starts:
    print("❌ Delivery middleware marker not found; unchanged.")
    sys.exit(2)
start = min(starts)

registration = "dp.callback_query.outer_middleware(_DfmtDeliveryMiddleware())"
reg_pos = src.find(registration, start)
if reg_pos < 0:
    print("❌ Delivery middleware registration not found; unchanged.")
    sys.exit(3)
end = src.find("\n", reg_pos + len(registration))
end = len(src) if end < 0 else end + 1
next_end = src.find("\n", end)
next_line = src[end:len(src) if next_end < 0 else next_end]
if "READY canonical" in next_line or "callback middleware active" in next_line:
    end = len(src) if next_end < 0 else next_end + 1

block = r'''# [DELIVERY_FINAL_V12]
# Canonical delivery: callback -> ownership -> archive recovery -> build -> upload.
# TXT is built inline (never falls back to the xlsx builder) and the format
# keyboard stays visible so the customer can re-download the same order.
from aiogram import BaseMiddleware as _DfmtBaseMiddleware


def _dlog(msg):
    print(f"[delivery-v12] {msg}", flush=True)


def _dtoken():
    for _name in ("BOT_TOKEN", "TOKEN", "API_TOKEN", "TG_TOKEN", "BOT_API_TOKEN"):
        _value = globals().get(_name)
        if isinstance(_value, str) and ":" in _value and len(_value) > 20:
            return _value
    _value = str(getattr(bot, "token", ""))
    if ":" in _value and len(_value) > 20:
        return _value
    raise RuntimeError("bot token not found")


async def _dupload_document(chat_id, filename, payload, caption, mime):
    import asyncio as _du_asyncio
    import json as _du_json
    import os as _du_os
    import re as _du_re
    import tempfile as _du_tempfile

    _safe = _du_re.sub(r"[^A-Za-z0-9._-]", "_", str(filename)) or "delivery.txt"
    _path = None
    _proc = None
    try:
        with _du_tempfile.NamedTemporaryFile(
                mode="wb", prefix="nx-delivery-", suffix="-" + _safe,
                dir="/tmp", delete=False) as _tmp:
            _tmp.write(payload)
            _path = _tmp.name

        _cmd = (
            "curl", "--ipv4", "--http1.1", "--no-keepalive",
            "--silent", "--show-error", "--fail-with-body",
            "--connect-timeout", "10", "--max-time", "50",
            "--request", "POST",
            "--form-string", f"chat_id={chat_id}",
            "--form-string", f"caption={caption or ''}",
            "--form", f"document=@{_path};filename={_safe};type={mime}",
            "--config", "-",
        )
        _config = f'url = "https://api.telegram.org/bot{_dtoken()}/sendDocument"\n'.encode()
        _proc = await _du_asyncio.create_subprocess_exec(
            *_cmd, stdin=_du_asyncio.subprocess.PIPE,
            stdout=_du_asyncio.subprocess.PIPE,
            stderr=_du_asyncio.subprocess.PIPE,
        )
        try:
            _out, _err = await _du_asyncio.wait_for(_proc.communicate(_config), 55)
        except _du_asyncio.TimeoutError:
            _proc.kill()
            await _proc.wait()
            raise TimeoutError("Telegram upload killed after 55 seconds")

        _raw = (_out or b"").decode("utf-8", "replace").strip()
        _error = (_err or b"").decode("utf-8", "replace").strip()
        if _proc.returncode != 0:
            raise RuntimeError(f"curl exit={_proc.returncode}: {(_error or _raw)[:300]}")
        try:
            _result = _du_json.loads(_raw)
        except Exception as _json_error:
            raise RuntimeError(f"Telegram returned invalid JSON: {_raw[:160]}") from _json_error
        if not _result.get("ok"):
            raise RuntimeError(f"Telegram rejected file: {_result.get('description', 'unknown')}")
        return True
    finally:
        if _proc is not None and _proc.returncode is None:
            try:
                _proc.kill()
                await _proc.wait()
            except ProcessLookupError:
                pass
        if _path:
            try:
                _du_os.unlink(_path)
            except FileNotFoundError:
                pass


def _dload_archive(sale_id):
    _cn = sqlite3.connect("/root/store.db", timeout=15)
    try:
        _cn.execute("PRAGMA busy_timeout=15000")
        return _cn.execute(
            "SELECT stock_id, data, category, user_id FROM delivery_archive "
            "WHERE sale_id=? ORDER BY id ASC", (sale_id,)
        ).fetchall()
    finally:
        _cn.close()


def _dsplit_account(raw):
    _text = (raw or "").replace("\r", " ").replace("\n", " ").strip()
    _parts = _text.split(None, 2)
    while len(_parts) < 3:
        _parts.append("")
    return _parts[0], _parts[1], _parts[2]


def _dbuild_txt(items, label, qty, sale_id):
    _lines = [
        f"NEXUS-X DELIVERY - {label}",
        f"Order #{sale_id} | Quantity: {qty}",
        "=" * 46,
        "",
    ]
    for _index, (_stock_id, _raw) in enumerate(items, start=1):
        _uid, _pwd, _cookies = _dsplit_account(_raw)
        _lines += [
            f"[{_index}] Stock ID: {_stock_id}",
            f"UID     : {_uid}",
            f"PASSWORD: {_pwd}",
            f"COOKIES : {_cookies}",
            "-" * 46,
            "",
        ]
    return ("\n".join(_lines)).encode("utf-8")


def _dbuild_xlsx(items, label, qty, sale_id):
    import io as _dx_io
    from openpyxl import Workbook as _DxWorkbook
    from openpyxl.styles import Font as _DxFont, PatternFill as _DxFill

    _wb = _DxWorkbook()
    _ws = _wb.active
    _ws.title = "Delivery"
    _ws["A1"] = f"NEXUS-X DELIVERY - {label} | Order #{sale_id} | Qty: {qty}"
    _ws["A1"].font = _DxFont(name="Arial", bold=True, size=12)
    _headers = ("#", "Stock ID", "UID", "PASSWORD", "COOKIES")
    for _col, _title in enumerate(_headers, start=1):
        _cell = _ws.cell(row=3, column=_col, value=_title)
        _cell.font = _DxFont(name="Arial", bold=True)
        _cell.fill = _DxFill("solid", start_color="FFFF00")
    for _row_index, (_stock_id, _raw) in enumerate(items, start=4):
        _uid, _pwd, _cookies = _dsplit_account(_raw)
        for _col, _value in enumerate((_row_index - 3, _stock_id, _uid, _pwd, _cookies), start=1):
            _cell = _ws.cell(row=_row_index, column=_col, value=_value)
            _cell.font = _DxFont(name="Arial")
    for _col_letter, _width in (("A", 6), ("B", 12), ("C", 26), ("D", 22), ("E", 90)):
        _ws.column_dimensions[_col_letter].width = _width
    _buffer = _dx_io.BytesIO()
    _wb.save(_buffer)
    return _buffer.getvalue()


class _DfmtDeliveryMiddleware(_DfmtBaseMiddleware):
    async def __call__(self, handler, event, data):
        import asyncio as _delivery_asyncio

        _cbdata = getattr(event, "data", "") or ""
        if not _cbdata.startswith("dfmt:"):
            return await handler(event, data)

        _sid = -1
        _dlog(f"click user={event.from_user.id} data={_cbdata}")
        try:
            _, _fmt, _sid_text = _cbdata.split(":", 2)
            if _fmt not in ("xlsx", "txt"):
                raise ValueError("invalid format")
            _sid = int(_sid_text)
        except Exception:
            await event.answer("Invalid delivery request", show_alert=True)
            return None

        try:
            try:
                await _delivery_asyncio.wait_for(
                    event.answer(f"{_fmt.upper()} তৈরি হচ্ছে..."), 15)
            except Exception:
                pass

            _meta = _PENDING_DELIVERY.get(_sid)
            if not _meta:
                _rows = await _delivery_asyncio.wait_for(
                    _delivery_asyncio.to_thread(_dload_archive, _sid), 25)
                if not _rows:
                    await event.answer("Delivery data পাওয়া যায়নি। Admin কে জানান।", show_alert=True)
                    return None
                _cat = _rows[0][2] or "item"
                _label = {
                    "fb61": "FB 61", "fb1000": "FB 1000", "tempid": "Temp ID",
                    "ig": "Instagram", "fb": "Facebook", "bmig": "BM IG", "bmfb": "BM FB",
                }.get(_cat, _cat.upper())
                _meta = {
                    "user_id": _rows[0][3], "cat": _cat, "lbl": _label,
                    "qty": len(_rows), "items": [(row[0], row[1]) for row in _rows],
                }
                _PENDING_DELIVERY[_sid] = _meta

            if _meta.get("user_id") and int(_meta["user_id"]) != int(event.from_user.id):
                await event.answer("এটা আপনার order নয়", show_alert=True)
                return None

            _items = _meta["items"]
            _label = _meta["lbl"]
            _qty = _meta["qty"]
            if _fmt == "xlsx":
                _payload = await _delivery_asyncio.wait_for(
                    _delivery_asyncio.to_thread(_dbuild_xlsx, _items, _label, _qty, _sid), 45)
                _filename = f"order-{_sid}-{_meta['cat']}.xlsx"
                _mime = ("application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet")
            else:
                _payload = await _delivery_asyncio.wait_for(
                    _delivery_asyncio.to_thread(_dbuild_txt, _items, _label, _qty, _sid), 30)
                _filename = f"order-{_sid}-{_meta['cat']}.txt"
                _mime = "text/plain"

            _dlog(f"file-ready sale={_sid} fmt={_fmt} name={_filename} bytes={len(_payload)}")
            _caption = f"📦 {_label} × {_qty} • Order #{_sid} ({_fmt.upper()})"
            await _dupload_document(
                event.message.chat.id, _filename, _payload, _caption, _mime)
            _dlog(f"sent sale={_sid} fmt={_fmt} name={_filename} bytes={len(_payload)}")
            # Keyboard is intentionally kept so the same order can be re-downloaded
            # in either format at any time.
        except Exception as _error:
            _dlog(f"error sale={_sid}: {type(_error).__name__}: {_error}")
            try:
                await _delivery_asyncio.wait_for(
                    event.message.answer("File পাঠানো যায়নি। আবার চেষ্টা করুন অথবা Admin কে জানান।"), 30)
            except Exception:
                pass
        return None


dp.callback_query.outer_middleware(_DfmtDeliveryMiddleware())
print("[delivery-v12] READY inline TXT/XLSX builders + re-download keyboard", flush=True)
'''

patched = src[:start] + block + src[end:]
required = (
    "# [DELIVERY_FINAL_V12]", "class _DfmtDeliveryMiddleware",
    "async def _dupload_document", "create_subprocess_exec",
    "def _dbuild_txt", "def _dbuild_xlsx", registration,
    "[delivery-v12] READY",
)
missing = [item for item in required if item not in patched]
if missing:
    print("❌ Generated block verification failed:", ", ".join(missing))
    sys.exit(4)
if patched.count(registration) != 1:
    print(f"❌ Found {patched.count(registration)} delivery registrations; unchanged.")
    sys.exit(5)
if "[delivery-v11] READY" in patched or "[delivery-v11]" in patched:
    print("❌ Old V11 markers still present; unchanged.")
    sys.exit(6)
try:
    compile(patched, STORE, "exec")
except SyntaxError as error:
    print(f"❌ Syntax error; unchanged: {error}")
    sys.exit(7)

backup = f"{STORE}.bak-delivery-v12-{int(time.time())}"
shutil.copy2(STORE, backup)
open(STORE, "w", encoding="utf-8").write(patched)
print(f"✅ Backup: {backup}")
print("✅ V12 installed: TXT built inline, XLSX built inline, correct MIME per format")
print("✅ Buttons kept after delivery -> same order re-download works any time")
